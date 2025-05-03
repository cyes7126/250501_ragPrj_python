import os
from dotenv import load_dotenv

import time
import pandas as pd
import fitz
import numpy as np
import logging
from logging.handlers import RotatingFileHandler
import inspect
import json
import subprocess
from bs4 import BeautifulSoup
from typing import List, Dict, Union
from logging.handlers import RotatingFileHandler
from langchain_huggingface import HuggingFaceEmbeddings
from langchain.retrievers import ContextualCompressionRetriever
from langchain.retrievers.document_compressors import CrossEncoderReranker
from langchain_community.retrievers import BM25Retriever
from langchain_community.retrievers import TFIDFRetriever
from langchain_community.cross_encoders import HuggingFaceCrossEncoder
from langchain.text_splitter import MarkdownHeaderTextSplitter
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_experimental.text_splitter import SemanticChunker
from CustomMilvus import CustomMilvus

from langchain.schema import Document
from fastapi import FastAPI, HTTPException
from fastapi.responses import JSONResponse
from fastapi.exceptions import RequestValidationError
from langchain_community.document_loaders import UnstructuredExcelLoader
from langchain_community.document_loaders import UnstructuredWordDocumentLoader
from langchain.retrievers import EnsembleRetriever
from langchain_community.document_loaders.csv_loader import CSVLoader
from langchain_community.document_loaders import UnstructuredHTMLLoader
from pydantic import BaseModel
import nest_asyncio
import uvicorn
from HWPLoader import HWPLoader
from HWPXLoader import Win32HwpLoader
import torch
import warnings
import re
from openpyxl import load_workbook
import html2text

# 페이지별 OCR
from PyPDF2 import PdfReader, PdfWriter
from tempfile import NamedTemporaryFile
import os
from io import BytesIO

# 마커 새로운버전
from marker.converters.pdf import PdfConverter
from marker.models import create_model_dict
from marker.output import text_from_rendered
from marker.config.parser import ConfigParser
from sympy.integrals.meijerint_doc import category

import random
import string
import math
import pdfplumber
from pdfminer.pdftypes import resolve1

# 최대 스레드 수 지정
import asyncio
max_thread_cnt = 30
semaphore = asyncio.Semaphore(max_thread_cnt)

warnings.filterwarnings("ignore", message=".*could benefit from vacuuming your database.*")

app = FastAPI()

model_name = "BAAI/bge-m3"  # 로컬 테스트용
#model_name = "/home/app/bge-m3"

# 설치가 필요한 cross-encoder 모델
reranker_name = "BAAI/bge-reranker-v2-m3" # 로컬 테스트용
#reranker_name = "/home/app/bge-reranker-v2-m3"

model_kwargs = {"device": "cuda"}
encode_kwargs = {"normalize_embeddings": True}

chunk_size = 500  # 텍스트 청크 사이즈
chunk_overlap = 20  # 청크 중복
chunk_k = 20  # 검색 결과 수
max_ocr_page_count = 5000
MAX_THREADS = os.cpu_count()  # CPU 코어 개수에 따라 동적으로 MAX_THREADS 설정, 시스템의 CPU 코어 개수로 설정
# 청크 데이터 관련 전역 변수
all_chunks = None
chunk_metadata = None

# RAG Name
rag_name = "gen_rag"  # lower case
reg_version = ""
# 설정
data_dir = f"data_{rag_name}{reg_version}"  # PDF 파일 디렉토리

# 기본 설정 값
if os.path.exists(".env"):
    load_dotenv(".env")
milvus_url = os.getenv("MILVUS_URL", "http://localhost:19530")
milvus_token = os.getenv("MILVUS_TOKEN", "minioadmin:minioadmin")
collection = os.getenv("CONTAINER_NAME", "default_container")

# Parsing 모델 로드하기
# model_list = load_all_models()

# 로거 초기화
logger = logging.getLogger(__name__)

# Marker 별도로 Convert 객체 생성
config_parser = ConfigParser({"force_ocr" : False, "language" : "en,ko"})
config_parser_ocr = ConfigParser({"force_ocr" : True, "language" : "en,ko"})
artifact_dict = create_model_dict()
converter = PdfConverter(
    config=config_parser.generate_config_dict(),
    artifact_dict=artifact_dict,
)
converter_ocr = PdfConverter(
    config=config_parser_ocr.generate_config_dict(),
    artifact_dict=artifact_dict,
)

# 병렬로 converter 객체 접근시 오류발생. 단일 접근 가능하도록 lock 처리
from threading import Lock
marker_lock = Lock()

# Encoding Checker
def is_valid_text(text, threshold=0.35, max_repeat=10, unique_threshold=0.3):
    """
    텍스트가 정상인지 판단합니다.
    1. 전체 텍스트 중 한글과 영문(대소문자) 비율이 threshold 미만이면 False.
    2. 어떤 문자가 max_repeat번 이상 연속되면 False.
    3. 전체 문자 대비 고유(유니크) 문자의 비율이 unique_threshold 미만이면 False.
    """
    if not text:
        return False

    # 1. 유효 문자 비율 계산 (한글, 영문)
    valid_chars = re.findall(r'[\uac00-\ud7a3a-zA-Z]', text)
    ratio = len(valid_chars) / len(text)
    if ratio < threshold:
        log(f"[DEBUG] 한글/영문 비율이 낮음 ({ratio:.2f}) - OCR 필요")
        return False

    # 2. 동일 문자가 max_repeat번 이상 연속되는 경우 체크
    if re.search(r'(.)\1{' + str(max_repeat - 1) + r',}', text):
        log(f"[DEBUG] 동일 문자 {max_repeat}회 이상 반복됨 - OCR 필요")
        return False
    
    # 무의미한 경우가 많아, 해당 부분 주석처리
    '''
    # 3. 고유 문자 비율 계산
    unique_chars = set(text)
    unique_ratio = len(unique_chars) / len(text)
    if unique_ratio < unique_threshold:
        log(f"[DEBUG] 고유 문자 비율이 낮음 ({unique_ratio:.2f}) - OCR 필요")
        return False
    '''
    return True

# pdf로 변환할 파일 확장자
pdf_convert_file_ext = (".ppt", ".pptx")

# Set up a rotating file handler and console handler
def initialize_logger():
    global logger
    logger.setLevel(logging.DEBUG)
    # 기존 핸들러 제거
    logger.handlers = []

    # File handler with rotating logs
    file_handler = RotatingFileHandler(
        filename='gen_rag.log',
        maxBytes=100 * 1024 * 1024,  # 100 MB
        backupCount=2,
        encoding='utf-8'
    )

    # Console handler
    console_handler = logging.StreamHandler()

    # Format for the log messages with detailed timestamp including milliseconds
    formatter = logging.Formatter(
        '[%(asctime)s.%(msecs)03d][%(levelname)s] %(message)s - %(custom_filename)s:%(custom_lineno)d',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    # Add handlers to the logger
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    # 루트 로거로의 전파 방지
    logger.propagate = False


initialize_logger()


# Logging function with added error handling
def log(*args, **kwargs):
    try:
        # 로그 메시지를 문자열로 변환
        log_message = ' '.join(map(str, args))

        # 호출자 정보 추출
        frame = inspect.currentframe().f_back
        file_name = os.path.basename(frame.f_code.co_filename)
        line_number = frame.f_lineno

        # 로그 메시지 기록 (extra 인자에 커스텀 키 사용)
        logger.debug(log_message, extra={'custom_filename': file_name, 'custom_lineno': line_number})

    except Exception as e:
        logger.error(f"[log] Error logging message: {e}")


# 임베딩 모델 로드하기
def load_embedding_model():
    log("[load Model Start]")
    # model = SentenceTransformer(model_name, device='cuda')
    model = HuggingFaceEmbeddings(
        model_name=model_name, model_kwargs=model_kwargs, encode_kwargs=encode_kwargs)
    log("[load Model End]")
    return model


# Milvus 데이터베이스 초기화 함수
def init_milvus():
    try:
        milvus_db = CustomMilvus(
            embedding_function=embedding_model,
            connection_args={"uri": milvus_url, "token": milvus_token},
            collection_name=collection,
            index_params={
                "index_type": "AUTOINDEX",
                "metric_type": "COSINE" # bge3는 cosine사용
            },
            consistency_level="Strong", # 일관성 레벨. 모든 노드가 최신 데이터를 봄
            drop_old=False, # False: 기존 컬렉션 유지
            auto_id=True, # 아이디 자동 생성
        )
        log(f"[Init Milvus Db] Success - {collection}")
        return milvus_db
    except Exception as e:
        log(f"[Init Milvus Db] Error: {e}")
        return None

def reset_milvus(milvus_db):
    if milvus_db is not None:
        milvus_db.drop_collection()
    milvus_db = init_milvus()
    return milvus_db


# 파일 내용 가공
def get_cleaned_text(text):
    cleaned_text = text
    return cleaned_text

# 허용된 thread 수만큼 제한
async def handle_limited_thread(func, *args, **kwargs):
    async with semaphore:
        return await asyncio.to_thread(func, *args, **kwargs)

# embedding 모델 로드하기
embedding_model = load_embedding_model()

# milvus 인스턴스 생성
milvus_db = init_milvus()

# 스플리터
headers_to_split_on = [("####", "Page Of Content")]
markdown_splitter = MarkdownHeaderTextSplitter(headers_to_split_on=headers_to_split_on)
recursive_splitter = RecursiveCharacterTextSplitter()
## 시맨틱 청킹 수행
splitter_semantic = SemanticChunker(embedding_model)
splitter_dictionary = {
    'Markdown': markdown_splitter,
    'Recursive': recursive_splitter,
    'Semantic': splitter_semantic
}

# 파일 정보를 page_content에 붙여, 양질의 검색 수행
def attach_file_info(file_name, page_number, text):
    if isinstance(text, Document):
        text = text.page_content
    return f"####Record ID: {file_name}, page Number: {page_number}\n\n" + text


# 한글(HWP) 파일에서 텍스트 추출
def extract_text_from_hwp(file_path: str) -> str:
    hwp_loader = HWPLoader(file_path)
    document = next(hwp_loader.lazy_load())
    return document.page_content


# PDF에서 텍스트 추출 + 실패시 OCR 로직
def extract_text_from_pdf(file_path, pdf_file_name, ocr_fallback=True):
    pdf_full_path = os.path.join(file_path, pdf_file_name)
    doc = fitz.open(pdf_full_path)
    ocr_target_pages = []
    text_by_page = {}

    for index in range(doc.page_count):
        page_num = index+1
        page = doc.load_page(index)
        text = page.get_text("text")
        text_by_page[page_num] = text
        if ocr_fallback and not is_valid_text(text):
            ocr_target_pages.append(page_num)

    # 인코딩 깨진 경우, Marker-pdf(OCR) 실행
    if ocr_fallback and len(ocr_target_pages) > 0:
        log(f"Pages {ocr_target_pages} 텍스트 깨짐 감지. OCR 재처리 실행.")
        use_memory = check_acl_permission(file_path)
        if use_memory:
            pdf_buffer = BytesIO()
        else:
            pdf_buffer = None

        tmp_file_name = make_tmp_pdf_file_select_pages(pdf_full_path, ocr_target_pages, use_memory, pdf_buffer)
        log(f"MarkerPdf(OCR)용 파일 생성 Page {ocr_target_pages} : {tmp_file_name}")
        results = extract_text_from_pdf_by_marker_pages(tmp_file_name, use_ocr=True)
        for r_idx, text in enumerate(results):
            page_number = ocr_target_pages[r_idx]
            text_by_page[page_number] = text

        os.remove(tmp_file_name)

    return text_by_page

# 사용할 스플리터 리턴
def get_splitter(recursive=False, splitter_name = ''):
    if splitter_name in splitter_dictionary:
        text_splitter = splitter_dictionary[splitter_name]
    else:
        # recursive가 참이면 recursive splitter를 사용하고 그렇지 않다면 markdown splitter를 사용
        text_splitter = recursive_splitter if recursive else splitter_semantic
    return text_splitter

# PDF page의 이미지가 유의미한지 체크
def is_meaningful_image(obj, page_width, page_height):
    render_w, render_h = obj["width"], obj["height"]
    filter_type = resolve1(obj["stream"].attrs.get("Filter"))
    colorspace = resolve1(obj.get("colorspace", []))
    byte_size = resolve1(obj["stream"].attrs.get("Length"))
    aspect_ratio = render_w / render_h if render_h else 0

    # render 이미지 크기가 50보다 작음
    if render_w < 50 or render_h < 50:
        return False

    # 세로에 비해 가로너비가 30배 이상일때 (분리 선일 확률 높음)
    if aspect_ratio > 30:
        return False

    # 압축된 바이트 크기
    # 1000 이하 : 로고 이미지, 5000 이하 : 간단한 그림, 배경, 도형일 수 있음
    if byte_size and byte_size < 1000:
        return False

    # 필터 타입 확인
    # FlateDecode : 단순 도형 or 선, DeviceGray : 단색선, 테두리 박스 등
    if isinstance(filter_type, list):
        filter_type = filter_type[0]
    if filter_type == "/FlateDecode" and "/DeviceGray" in colorspace:
        return False

    # 이미지의 시작위치가 -인지 확인
    if obj['x0'] < 0 or obj['y0'] < 0:
        return False
    
    # 이미지의 크기가 페이지 전체 크기를 초과하는지 확인
    if obj['x1'] > page_width or obj['y1'] > page_height:
        return False

    return True

# PDF에서 이미지, 테이블 존재 여부 확인 && 텍스트 추출
def analyze_pdf_range_structure(pdf_path, start_page, end_page, ocr_fallback=True):
    def analyze_page(plumber_page, page_width, page_height):
        content = plumber_page.extract_text()
        # 추출한 텍스트의 인코딩이 깨지면 has_image로 처리
        if ocr_fallback and not is_valid_text(content):
            return "has_image"

        # 유효한 이미지 존재 여부
        page_images = plumber_page.objects.get("image")
        has_image = False
        if page_images is not None:
            valid_images = [img for img in page_images if is_meaningful_image(img, page_width, page_height)]
            if len(valid_images) > 0:
                has_image = True

        # 테이블 존재 여부
        has_table = bool(plumber_page.extract_table())

        if has_image:
            content = "has_image"
        elif has_table:
            content = "has_table"
        return content

    layout_by_page = {}
    with pdfplumber.open(pdf_path) as pdf:
        page_width = pdf.pages[0].width
        page_height = pdf.pages[0].height

        for i in range(start_page, end_page + 1):
            plumber_page = pdf.pages[i]
            layout_by_page[i + 1] = analyze_page(plumber_page, page_width, page_height)
    return layout_by_page

# 스플리터에 따라 텍스트 분할
def chunk_text(directory, file_save_name, file_name, page_num, text, text_splitter):
    filename, ext = os.path.splitext(os.path.basename(file_save_name))
    chunks = text_splitter.split_text(text)
    chunked_texts = []

    # 청킹이 수행된 각 텍스트의 인덱스에 위에 설정한 파일명, 페이지 번호, 청킹 결과
    for chunk in chunks:
        chunked_texts.append(attach_file_info(filename, page_num + 1, chunk))
    return chunked_texts

# 텍스트를 청크 사이즈 기준으로 자름
def split_text_by_chunk_size(text, chunk_size=chunk_size, chunk_overlap=chunk_overlap):
    step = chunk_size - chunk_overlap
    words = text.split()
    all_chunks = [
        ' '.join(words[i:i+chunk_size])
        for i in range(0, len(words), step)
    ]
    return all_chunks

def extract_text_from_file(file_path, file_save_name, use_marker=True, ocr_fallback=True):
    text_by_page = {}
    # 엑셀 파일(xlsx, xls) 처리
    if file_save_name.lower().endswith((".xlsx", ".xls")):
        wb = load_workbook(filename=os.path.join(file_path, file_save_name), read_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.rows)
            header_index = 0

            for idx, row in enumerate(rows):
                if any(cell.value is not None for cell in row):
                    header_index = idx
                    break

            headers = [cell.value for cell in rows[header_index]]
            row_strings = []

            for row in rows[header_index + 1:]:
                if not any(cell.value for cell in row):
                    continue
                row_string = ""
                for header, cell in zip(headers, row):
                    value = str(cell.value) if cell.value is not None else ''
                    if value:
                        row_string += f"{header} : {value}\n"

                if row_string.strip():
                    row_strings.append(row_string)
            text_by_page[1] = row_strings
        wb.close()

    # CSV 파일 처리
    elif file_save_name.lower().endswith(".csv"):
        loader = CSVLoader(file_path=os.path.join(file_path, file_save_name))
        docs = loader.load()
        text_by_page[1] = [doc.page_content for doc in docs]

    elif file_save_name.lower().endswith('.hwp'):
        try:
            # 기존 HWPLoader를 이용한 HWP 텍스트 추출 시도
            result_text = extract_text_from_hwp(os.path.join(file_path, file_save_name))
            text_by_page[1] = result_text
        except Exception as e:
            # 오류 메시지에 "OLE2"가 포함된 경우 PDF 변환 후 파싱 진행
            if "OLE2" in str(e):
                log(f"[extract_text_from_file] HWP 파싱 오류 발생: {e}. PDF 변환 후 파싱 시도 중...")
                pdf_file_name = file_save_name
                # pdf가 아닌 경우, pdf로 convert
                if not file_save_name.lower().endswith('.pdf'):
                    convert_file_to_pdf(output_dir=file_path, target_file=os.path.join(file_path, file_save_name))
                    pdf_file_name = os.path.splitext(file_save_name)[0] + ".pdf"

                pdf_file_length = get_pdf_page_count(os.path.join(file_path, pdf_file_name))

                if use_marker:
                    text_by_page = preprocess_pdf_text_by_page_range(file_path, pdf_file_name, 0, pdf_file_length -1, ocr_fallback)
                else:
                    text_by_page = extract_text_from_pdf(file_path, pdf_file_name, ocr_fallback)

                try:
                    os.remove(os.path.join(file_path, pdf_file_name))
                    log(f"[extract_text_from_file] HWP -> 임시 PDF 파일 삭제 완료: {pdf_file_name}")
                except OSError as del_err:
                    log(f"[extract_text_from_file] HWP -> 임시 PDF 파일 삭제 중 오류 발생: {pdf_file_name}, {del_err}")
            else:
                raise e

    elif file_save_name.lower().endswith(".hwpx"):
        loader = Win32HwpLoader(os.path.join(file_path, file_save_name))
        docs = loader.load()
        text_by_page[1] = [doc.page_content for doc in docs]

    elif file_save_name.lower().endswith(".htm") or file_save_name.lower().endswith(".html"):
        loader = UnstructuredHTMLLoader(os.path.join(file_path, file_save_name), mode="elements")
        docs = loader.load()
        # html tag가 있는경우 markdown형태로 page_content update
        docs = convert_html_to_markdown(docs)
        text_by_page[1] = [doc.page_content for doc in docs]

    elif file_save_name.lower().endswith(".doc") or file_save_name.lower().endswith(".docx"):
        loader = UnstructuredWordDocumentLoader(os.path.join(file_path, file_save_name), mode="elements")
        docs = loader.load()
        # html tag가 있는경우 markdown형태로 page_content update
        docs = convert_html_to_markdown(docs)
        text_by_page[1] = [doc.page_content for doc in docs]

    # TXT
    elif file_save_name.lower().endswith(".txt"):
        with open(os.path.join(file_path, file_save_name), 'r', encoding='utf-8') as f:
            text = f.read()
        text_by_page[1] = split_text_by_chunk_size(text)

    return text_by_page


# 파싱 + 청킹 수행 함수
def extract_and_chunk_text_from_file(file_path, file_save_name, file_name, recursive, use_marker=True, category_pk=None, splitter_name="", ocr_fallback=True):
    all_chunks = []
    chunk_metadata = []
    parsed_text = ''
    text_splitter = get_splitter(recursive, splitter_name)
    log(f"텍스트 추출 시작. splitter : {type(text_splitter).__name__}")

    file_size = os.path.getsize(os.path.join(file_path, file_save_name))
    if file_size == 0:
        raise HTTPException(status_code=400, detail={
            "error": {
                "description": "Empty file"
            }
        })

    def process_text_by_page(file_path, file_save_name, file_name, text_by_page, all_chunks, chunk_metadata):
        nonlocal parsed_text  # 외부 함수의 변수를 사용하기 위해 nonlocal 선언
        for page_num, text in enumerate(text_by_page):
            cleaned_text = get_cleaned_text(text)
            parsed_text += cleaned_text
            chunks = chunk_text(file_path, file_save_name, file_name, page_num, cleaned_text, text_splitter=text_splitter)
            for chunk_number, chunk in enumerate(chunks):
                all_chunks.append(chunk)
                chunk_metadata.append({
                    "file_path": file_path,
                    "file_save_name": file_save_name,
                    "file_name": file_name,
                    "page_number": page_num + 1,
                    "chunk_number": chunk_number + 1,
                    "text": chunk,
                    "category_pk": category_pk
                })

    # pdf로 변환할 파일 확장자일 경우
    if file_save_name.lower().endswith(pdf_convert_file_ext):
        convert_file_to_pdf(output_dir=file_path, target_file=os.path.join(file_path, file_save_name))
        pdf_file_name = os.path.splitext(file_save_name)[0] + ".pdf"

        # pdf에서 텍스트 추출
        text_by_page = preprocess_pdf_text_by_page(file_path, pdf_file_name, pdf_file_name, use_marker, ocr_fallback)
        process_text_by_page(file_path, file_save_name, file_name, text_by_page, all_chunks, chunk_metadata)

        # PDF 파일 삭제
        try:
            os.remove(os.path.join(file_path, pdf_file_name))
            log(f"[extract_and_chunk_text_from_file] Deleted PDF file: {pdf_file_name}")
        except OSError as e:
            log(f"[extract_and_chunk_text_from_file] Error deleting file: {pdf_file_name}, {e}")

    # PDF
    elif file_save_name.lower().endswith(".pdf"):
        text_by_page = preprocess_pdf_text_by_page(file_path, file_save_name, file_name, use_marker)
        process_text_by_page(file_path, file_save_name, file_name, text_by_page, all_chunks, chunk_metadata)

    # 엑셀 파일(xlsx, xls) 처리
    elif file_save_name.lower().endswith((".xlsx", ".xls")):
        wb = load_workbook(filename=os.path.join(file_path, file_save_name), read_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.rows)
            header_index = 0
            
            for idx, row in enumerate(rows):
                if any(cell.value is not None for cell in row):
                    header_index = idx
                    break

            headers = [cell.value for cell in rows[header_index]]
            row_strings = []

            for row in rows[header_index + 1:]:
                if not any(cell.value for cell in row):
                    continue
                row_string = ""
                for header, cell in zip(headers, row):
                    value = str(cell.value) if cell.value is not None else ''
                    if value:
                        row_string += f"{header} : {value}\n"

                if row_string.strip():
                    row_strings.append(row_string)


            for cell_index, cell in enumerate(row_strings):
                all_chunks.append(cell)
                chunk_metadata.append({
                    "file_path": file_path,
                    "file_save_name": file_save_name,
                    "file_name": file_name,
                    "page_number": 1,
                    "chunk_number": cell_index + 1,
                    "text": cell,
                    "category_pk": category_pk
                })
        wb.close()

    # CSV 파일 처리
    elif file_save_name.lower().endswith(".csv"):
        loader = CSVLoader(file_path=os.path.join(file_path, file_save_name))
        docs = loader.load()
        parsed_text = "\n".join([doc.page_content for doc in docs])
        for doc_index, doc in enumerate(docs):
            all_chunks.append(doc.page_content)
            chunk_metadata.append({
                "file_path": file_path,
                "file_save_name": file_save_name,
                "file_name": file_name,
                "page_number": 1,
                "chunk_number": doc_index + 1,
                "text": doc.page_content,
                "category_pk": category_pk
            })

    elif file_save_name.lower().endswith('.hwp'):
        try:
            # 기존 HWPLoader를 이용한 HWP 텍스트 추출 시도
            result_text = extract_text_from_hwp(os.path.join(file_path, file_save_name))
            parsed_text = result_text
            chunks = chunk_text(file_path, file_save_name, file_name, 0, result_text, text_splitter=text_splitter)
            for chunk_number, chunk in enumerate(chunks):
                all_chunks.append(chunk)
                chunk_metadata.append({
                    "file_path": file_path,
                    "file_save_name": file_save_name,
                    "file_name": file_name,
                    "page_number": 1,  # HWP 파일은 페이지 정보가 없으므로 1로 설정
                    "chunk_number": chunk_number + 1,
                    "text": chunk,
                    "category_pk": category_pk
                })
                
        except Exception as e:
            # 오류 메시지에 "OLE2"가 포함된 경우 PDF 변환 후 파싱 진행
            if "OLE2" in str(e):
                log(f"[extract_and_chunk_text_from_file] HWP 파싱 오류 발생: {e}. PDF 변환 후 파싱 시도 중...")
                pdf_file_name = os.path.splitext(file_save_name)[0] + ".pdf"
                subprocess.run([
                    "soffice", "--headless", "--convert-to", "pdf",
                    "--outdir", os.path.join(file_path), os.path.join(file_path, file_save_name)
                ], check=True)
                # PDF 전처리 함수를 이용해 페이지별 텍스트 추출
                text_by_page = preprocess_pdf_text_by_page(file_path, pdf_file_name, pdf_file_name, use_marker)
                # PDF 전용 파싱 함수로 청킹 수행 (이 부분은 PDF 파싱 시 사용하는 방식 그대로)
                process_text_by_page(file_path, file_save_name, file_name, text_by_page, all_chunks, chunk_metadata)
                parsed_text = "\n".join(text_by_page)
                try:
                    os.remove(os.path.join(file_path, pdf_file_name))
                    log(f"[extract_and_chunk_text_from_file] 임시 PDF 파일 삭제 완료: {pdf_file_name}")
                except OSError as del_err:
                    log(f"[extract_and_chunk_text_from_file] PDF 파일 삭제 중 오류 발생: {pdf_file_name}, {del_err}")
            else:
                raise e

    elif file_save_name.lower().endswith(".hwpx"):
        loader = Win32HwpLoader(os.path.join(file_path, file_save_name))
        docs = loader.load()
        parsed_text = "\n".join([doc.page_content for doc in docs])
        for doc_index, doc in enumerate(docs):
            all_chunks.append(doc.page_content)
            chunk_metadata.append({
                "file_path": file_path,
                "file_save_name": file_save_name,
                "file_name": file_name,
                "page_number": 1,
                "chunk_number": doc_index + 1,
                "text": doc.page_content,
                "category_pk": category_pk
            })

    elif file_save_name.lower().endswith(".htm") or file_save_name.lower().endswith(".html"):
        loader = UnstructuredHTMLLoader(os.path.join(file_path, file_save_name), mode="elements")
        docs = loader.load()
        # html tag가 있는경우 markdown형태로 page_content update
        docs = convert_html_to_markdown(docs)
        parsed_text = "\n\n".join([doc.page_content for doc in docs])
        combine_doc = docs[0]
        combine_doc.page_content = parsed_text
        all_chunks.append(combine_doc)
        chunk_metadata.append({
            "file_path": file_path,
            "file_save_name": file_save_name,
            "file_name": file_name,
            "page_number": 1,  # 페이지는 무조건 1
            "chunk_number": 1,
            "text": parsed_text,
            "category_pk": category_pk
        })

    elif file_save_name.lower().endswith(".doc") or file_save_name.lower().endswith(".docx"):
        loader = UnstructuredWordDocumentLoader(os.path.join(file_path, file_save_name), mode="elements")
        docs = loader.load()

        # html tag가 있는경우 markdown형태로 page_content update
        docs = convert_html_to_markdown(docs)
        parsed_text = "\n\n".join([doc.page_content for doc in docs])

        combine_doc = docs[0]
        combine_doc.page_content = parsed_text
        all_chunks.append(combine_doc)
        chunk_metadata.append({
            "file_path": file_path,
            "file_save_name": file_save_name,
            "file_name": file_name,
            "page_number": 1,  # 페이지는 무조건 1
            "chunk_number": 1,
            "text": parsed_text,
            "category_pk": category_pk
        })

    # TXT
    elif file_save_name.lower().endswith(".txt"):
        with open(os.path.join(file_path, file_save_name), 'r', encoding='utf-8') as f:
            text = f.read()
        parsed_text += text
        all_chunks = split_text_by_chunk_size(text)
        chunk_metadata = [
            {
                "file_path": file_path,
                "file_save_name": file_save_name,
                "file_name": file_name,
                "page_number": 1,  # TXT 파일에는 페이지 개념이 없으므로 1로 설정
                "chunk_number": i + 1,
                "text": chunk,
                "category_pk": category_pk
            }
            for i, chunk in enumerate(all_chunks)
        ]

    return parsed_text, chunk_metadata

def check_acl_permission(file_path):
    """
    ls -ld 명령어 결과의 10번째 문자가 '+'이면 ACL이 설정되었다고 판단합니다.
    ACL이 설정되어 있으면 True, 그렇지 않으면 False를 반환합니다.
    """
    try:
        ls_output = subprocess.check_output(['ls', '-ld', file_path]).decode('utf-8').strip()
        # ls 출력의 권한 문자열은 예: "drwxr-xr-x+" 로, 인덱스 10의 문자가 ACL 여부를 나타냅니다.
        if ls_output[10] == '+':
            print(f"ACL permissions are set for: {file_path}")
            return True
        else:
            print(f"No ACL permissions set for: {file_path}")
            return False
    except subprocess.CalledProcessError as e:
        print(f"Error CalledProcessError for {file_path}: {e}")
        return False
    except Exception as e:
        print(f"Error for {file_path}: {e}")
        return False


# Marker Pdf 실행된 text 반환
def extract_text_from_pdf_by_marker(pdf_file_name, use_ocr=False):
    try:
        if use_ocr:
            selected_converter = converter_ocr
        else:
            selected_converter = converter
        rendered = selected_converter(pdf_file_name)
        page_text, _, _ = text_from_rendered(rendered)
        return page_text[0] if page_text and len(page_text) > 0 else ""
    except Exception as e:
        raise HTTPException(status_code=400, detail={
            "error": {
                "description": f"Failed to extract pdf by marker({pdf_file_name}) : {e}"
            }
        })

# 여러페이지가 있는 pdf 파일
def extract_text_from_pdf_by_marker_pages(pdf_file_name, use_ocr=False):
    log(f"extract_text_from_pdf_by_marker_pages : {pdf_file_name}")

    if not os.path.exists(pdf_file_name) or os.path.getsize(pdf_file_name) == 0:
        raise HTTPException(status_code=400, detail={
            "error": {
                "description": f"Invalid temp PDF file ({pdf_file_name})"
            }
        })

    # convert 객체 동시 접근시 torch shape mismatch 발생
    with marker_lock:
        try:
            if use_ocr:
                selected_converter = converter_ocr
            else:
                selected_converter = converter
            rendered = selected_converter(pdf_file_name)
            page_text, _, _ = text_from_rendered(rendered)
            return page_text if page_text and len(page_text) > 0 else []
        except Exception as e:
            raise HTTPException(status_code=400, detail={
                "error": {
                    "description": f"Failed to extract pdf by marker({pdf_file_name}) : {e}"
                }
            })

def make_tmp_pdf_file(file_path, file_full_path, page_index, use_memory=False, pdf_buffer=None):
    reader = PdfReader(file_full_path)
    try:
        # 메모리 기반 방식
        if use_memory:
            writer = PdfWriter()
            writer.add_page(reader.pages[page_index])
            writer.write(pdf_buffer)
            pdf_buffer.seek(0)

            with NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_pdf:
                tmp_pdf.write(pdf_buffer.getvalue())
                tmp_pdf.flush()  # 디스크에 반영
                return tmp_pdf.name

        else: # 디스크 기반 방식
            writer = PdfWriter()
            writer.add_page(reader.pages[page_index])
            temp_page_file = os.path.join(file_path, f"temp_page_{page_index + 1}.pdf")
            with open(temp_page_file, "wb") as f:
                writer.write(f)
            return temp_page_file
    except Exception as e:
        raise HTTPException(status_code=400, detail={
            "error": {
                "description": f"Failed to make tmp_pdf_file {file_full_path} {page_index+ 1 } Page : {e}"
            }
        })



# 랜덤 문자열 생성
def generate_random_string(length=5):
    chars = string.ascii_letters + string.digits  # 알파벳 대소문자 + 숫자
    return ''.join(random.choices(chars, k=length))

# 고유한 임시 파일명 생성
def generate_unique_temp_filename(original_path, random_length=3):
    base, ext = os.path.splitext(original_path)

    while True:
        random_part = generate_random_string(random_length)
        temp_file = f"{base}_{random_part}{ext}"
        if not os.path.exists(temp_file):
            return temp_file  # 파일이 존재하지 않으면 반환


# 원하는 페이지만 pdf 파일 새로 생성
def make_tmp_pdf_file_select_pages(file_full_path, pages = [], use_memory=False, pdf_buffer=None):
    # 페이지 번호 없는 경우
    if(len(pages) == 0):
        raise HTTPException(status_code=400, detail={
            "error": {
                "description": f"Failed to make tmp_pdf_file {file_full_path} - No Pages"
            }
        })

    reader = PdfReader(file_full_path)
    writer = PdfWriter()
    for page_number in pages:
        # 해당 페이지 인덱스 추가
        writer.add_page(reader.pages[page_number - 1])

    try:
        # 메모리 기반 방식
        if use_memory:
            writer.write(pdf_buffer)
            pdf_buffer.seek(0)
            tmp_file_name = ''

            with NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_pdf:
                tmp_pdf.write(pdf_buffer.getvalue())
                tmp_pdf.flush()  # 디스크에 반영
                tmp_file_name = tmp_pdf.name
            return tmp_file_name

        else: # 디스크 기반 방식
            # 고유한 파일명 생성
            temp_page_file = generate_unique_temp_filename(file_full_path, 3)
            with open(temp_page_file, "wb") as f:
                writer.write(f)
            return temp_page_file
    except Exception as e:
        raise HTTPException(status_code=400, detail={
            "error": {
                "description": f"Failed to make tmp_pdf_file {file_full_path} {pages} Page : {e}"
            }
        })

# PDF 파일의 전처리 여부를 받아, text return
def preprocess_pdf_text_by_page(file_path, file_save_name, file_name, use_marker, ocr_fallback=True):
    pdf_full_path = os.path.join(file_path, file_save_name)
    
    if use_marker:
        # ACL 여부에 따라 메모리 기반 방식과 디스크 기반 방식 분기
        if check_acl_permission(file_path):
            # ACL이 설정되어 있으면 메모리 기반 방식 사용
            log("use_marker True 및 ACL 권한이 설정되어 있음. 메모리 기반 임시 파일 생성 방식을 사용합니다.")
            reader = PdfReader(pdf_full_path)
            num_pages = len(reader.pages)
            text_by_page = []
            broken_page_indices = []
            
            for i in range(num_pages):
                # BytesIO를 사용하여 메모리 내 PDF 생성
                pdf_buffer = BytesIO()
                writer = PdfWriter()
                writer.add_page(reader.pages[i])
                writer.write(pdf_buffer)
                pdf_buffer.seek(0)
                
                # NamedTemporaryFile을 사용해 임시 PDF 파일 생성
                with NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_pdf:
                    tmp_pdf.write(pdf_buffer.getvalue())
                    tmp_pdf.flush()  # 디스크에 반영
                    
                    # OCR 처리 (기본적으로 force_ocr=False)
                    content = extract_text_from_pdf_by_marker(tmp_pdf.name, False)
                    
                    # OCR 후 텍스트 유효성 검사 및 재처리
                    if ocr_fallback and not is_valid_text(content):
                        log(f"Page {i+1} 텍스트 깨짐 감지. OCR 재처리 실행.")
                        broken_page_indices.append(i)
                        content = extract_text_from_pdf_by_marker(tmp_pdf.name, True)

                os.remove(tmp_pdf.name)
                text_by_page.append(content)
                pdf_buffer.close()
            
            log(f"깨진 페이지 인덱스 (0-based): {broken_page_indices}")
            return text_by_page
        
        else:
            # ACL이 설정되어 있지 않으면 디스크 기반 방식 사용
            log("use_marker True이나 ACL 권한이 설정되어 있지 않음. 파일 시스템 기반 임시 PDF 파일 생성 방식을 사용합니다.")
            reader = PdfReader(pdf_full_path)
            num_pages = len(reader.pages)
            text_by_page = []
            temp_page_files = []
            
            for i in range(num_pages):
                writer = PdfWriter()
                writer.add_page(reader.pages[i])
                temp_page_file = os.path.join(file_path, f"temp_page_{i+1}.pdf")
                with open(temp_page_file, "wb") as f:
                    writer.write(f)
                temp_page_files.append(temp_page_file)
                
                content = extract_text_from_pdf_by_marker(temp_page_file, False)

                if ocr_fallback and not is_valid_text(content):
                    log(f"Page {i+1} 텍스트 깨짐 감지. OCR 재처리 실행.")
                    content = extract_text_from_pdf_by_marker(temp_page_file, True)
                
                text_by_page.append(content)
                log(f'문서 파싱 결과 : {content}')
            
            # 생성된 임시 파일 삭제
            for temp_file in temp_page_files:
                os.remove(temp_file)
            return text_by_page
    
    # use_marker가 False인 경우엔 기본 텍스트 추출 방식을 사용
    else:
        text_by_page = extract_text_from_pdf(file_path, file_save_name, False)
        return list(text_by_page.values())


# 청크 데이터를 Document 객체로 변환하는 함수 추가
def convert_to_documents(chunked_data, category):
    documents = []

    for chunk in chunked_data[1]:  # chunked_data[1]은 메타데이터 리스트
        # 청크 문서
        text = chunk["text"].encode("utf-8", errors="replace").decode("utf-8")

        # 메타데이터 정보 설정
        metadata = {
            "file_path": chunk["file_path"],
            "file_save_name": chunk["file_save_name"],
            "file_name": chunk["file_name"],
            "page_number": chunk["page_number"],
            "chunk_number": chunk["chunk_number"],
            "category": category,
            "category_pk": chunk["category_pk"],
            "reg_dt": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time())),
            "additional": "" # 추가적인 정보
        }

        # Document 생성 (page_content는 text로 설정)
        document = Document(page_content=text, metadata=metadata)
        documents.append(document)

    return documents


def load_or_create_embeddings(file_path, file_save_name, file_name, category, recursive, use_marker, category_pk,
                              collection_name=collection, splitter_name=""):
    if milvus_db is None:
        log("[load_or_create_embeddings] Milvus DB instance not found")
        raise HTTPException(status_code=404, detail="Milvus DB instance not found")

    try:
        chunk_metadata = extract_and_chunk_text_from_file(file_path, file_save_name, file_name, recursive, use_marker, category_pk, splitter_name=splitter_name)
        
        if not chunk_metadata[1]:
            raise HTTPException(status_code=400, detail={
                "error": {
                    "description": "No content could be extracted from document"
                }
            })

        mark_documents = convert_to_documents(chunk_metadata, category)
        milvus_db.add_documents(mark_documents)

    except HTTPException as he:
        # HTTPException은 그대로 전달
        raise he
        
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail={
            "error": {
                "description": f"Failed to process document: {str(e)}"
            }
        })

def get_full_chunk_text_list(category):
    expr = f'category_pk == "{category}"'
    results = milvus_db.search_by_metadata(expr=expr)
    sorted_results = sorted(results, key=lambda doc: (doc.metadata["file_save_name"], doc.metadata["page_number"], doc.metadata["chunk_number"]))
    selected_data_dict = {}

    for doc in sorted_results:
        file_save_name = doc.metadata["file_save_name"]
        if not file_save_name in selected_data_dict:
            selected_data_dict[file_save_name] = {
                "file_save_name": file_save_name,
                'file_path': doc.metadata['file_path'],
                'full_text': doc.page_content
            }
        else:
            selected_data_dict[file_save_name]['full_text'] += "\n\n\n"+doc.page_content
    return list(selected_data_dict.values())


def get_full_chunk_text(categoryPk:str, file_save_name:str):
    expr = f'category_pk == "{categoryPk}" and file_save_name == "{file_save_name}"'
    results = milvus_db.search_by_metadata(expr=expr)
    sorted_results = sorted(results, key=lambda doc: (
    doc.metadata["file_save_name"], doc.metadata["page_number"], doc.metadata["chunk_number"]))

    selected_data_dict = {
        'full_text': "\n\n\n".join([doc.page_content for doc in sorted_results]),
        'file_path': sorted_results[0].metadata['file_path'],
        'file_save_name': sorted_results[0].metadata['file_save_name'],
        'file_name': sorted_results[0].metadata['file_name'],
    }
    log("[full_text_metadata] text \n\n", selected_data_dict['full_text'])
    return selected_data_dict

# Retriever 설정 함수 수정
def get_reranker(file_save_name_list=[], top_n=chunk_k, category_pk=[]):
    expr = ""
    if len(category_pk) > 0:
        category_in = '","'.join(category_pk)
        expr += f'category_pk in ["{category_in}"]'

    if len(file_save_name_list) > 0:
        # 파일명만 추출하여 where절 생성
        only_file_save_name_list = [str.split("/")[-1] for str in file_save_name_list]
        file_name_in = '","'.join(only_file_save_name_list)
        file_expr = f'file_save_name in ["{file_name_in}"]'
        if expr:
            expr += " or " + file_expr
        else:
            expr = file_expr

    log(f"[get_reranker] 필터 조건 expr : {expr}")

    try:
        docs = milvus_db.search_by_metadata(expr=expr)
        fetch_k = top_n * 2

        # 리트리버 설정
        milvus_retriever = milvus_db.as_retriever(search_type="mmr", search_kwargs={"k": top_n, "fetch_k": fetch_k, "expr": expr})
        bm25_retriever = BM25Retriever.from_documents(docs)
        ensenble_retriver = EnsembleRetriever(retrievers=[bm25_retriever, milvus_retriever], weights=[0.6, 0.4])

        # 리랭커 모델 로드
        model = HuggingFaceCrossEncoder(model_name=reranker_name)
        compressor = CrossEncoderReranker(model=model, top_n=top_n)

        # 문서 압축 검색기 초기화
        compression_retriever = ContextualCompressionRetriever(base_compressor=compressor, base_retriever=ensenble_retriver)
        return compression_retriever
    except Exception as e:
        log(f"[get_reranker] Exception : \n {e}")
        raise e

def keyword_search(category_list=[], file_save_name_list=[], top_k=chunk_k):
    expr = ""
    if len(category_list) > 0:
        category_in = '","'.join(category_list)
        expr += f'category_pk in ["{category_in}"]'

    if len(file_save_name_list) > 0:
        file_name_in = '","'.join(file_save_name_list)
        file_expr = f'file_save_name in ["{file_name_in}"]'
        if expr:
            expr += " or " + file_expr
        else:
            expr = file_expr

    log(f"[keyword_search] 필터 조건 expr : {expr}")

    try:
        docs = milvus_db.search_by_metadata(expr=expr)
        # 리트리버 설정
        bm25_retriever = BM25Retriever.from_documents(docs)
        bm25_retriever.k = top_k
    #    tfidf_retriever = TFIDFRetriever.from_documents(bm25_converted_data)
    #    tfidf_retriever.k = top_k
    #    ensenble_retriver = EnsembleRetriever(retrievers=[bm25_retriever, tfidf_retriever], weights=[0.5, 0.5])
        return bm25_retriever
    except Exception as e:
        log(f"[keyword_search] Exception : \n {e}")
        raise e

# JSON 변환 함수
def convert_to_serializable(data):
    if isinstance(data, list):
        return [
            {
                "pageContent": doc.page_content,
                "page": doc.metadata.get("page_number"),
                "source": f"{doc.metadata.get('file_name')} {str(doc.metadata.get('page_number'))}page",
                "file_path": doc.metadata.get("file_path"),
                "file_save_name": doc.metadata.get("file_save_name"),
                "file_name": doc.metadata.get("file_name"),
                "chunk_number": doc.metadata.get("chunk_number"),
                "category_pk": doc.metadata.get("category_pk"),
                "reg_dt": doc.metadata.get("reg_dt", ""),
                "additional": doc.metadata.get("additional", "")
            }
            for doc in data
        ]
    return data

# Document의 html 태그를 markdown형태로 변경함
def convert_html_to_markdown(docs):
    for doc in docs:
        if 'text_as_html' in doc.metadata:
            markdown_content = html2text.html2text(doc.metadata['text_as_html'])
            if 'table' in doc.metadata['text_as_html']:
                markdown_content = "### 표 데이터\n" + markdown_content
            doc.page_content = markdown_content
    return docs


# 문서 삭제 함수
def delete_document_from_milvus(file_save_name: str):
    if milvus_db is None:
        log("[delete_document_from_milvus] Milvus DB instance not found")
        raise HTTPException(status_code=404, detail="Milvus DB instance not found")

    expr = f'file_save_name == "{file_save_name}"'
    results = milvus_db.search_by_metadata(expr=expr, limit=1)
    if len(results) == 0:
        log("[delete_document_from_milvus] Milvus DB Search Result not found")
        raise HTTPException(status_code=404, detail={
            "error": {
                "description": "Document not found in metadata"
            }
        })
    try:
        # 해당 문서의 모든 청크 삭제
        milvus_db.delete(expr=expr)
    except Exception as e:
        log(f"[delete_document_from_milvus] Milvus DB delete Error : {e}")
        raise HTTPException(status_code=400, detail={
            "error": {
                "description": f"Milvus DB delete Error : {e}"
            }
        })
    return JSONResponse({"documentResult": {"result": True, "status_code": "success",
                                            "description": "Document vectors deleted successfully"}}, status_code=200)


def delete_document_from_milvus_by_session_id(session_id: str):
    if milvus_db is None:
        log("[delete_document_from_milvus] Milvus DB instance not found")
        raise HTTPException(status_code=404, detail="Milvus DB instance not found")
    expr = f'file_path == "{session_id}"'
    results = milvus_db.search_by_metadata(expr=expr, limit=1)
    if len(results) == 0:
        log("[delete_document_from_milvus] Milvus DB Search Result not found")
        raise HTTPException(status_code=404, detail={
            "error": {
                "description": "Document not found in metadata"
            }
        })

    try:
        # 해당 문서의 모든 청크 삭제
        milvus_db.delete(expr=expr)
    except Exception as e:
        log(f"[delete_document_from_milvus] Milvus DB delete Error : {e}")
        raise HTTPException(status_code=400, detail={
            "error": {
                "description": f"Milvus DB delete Error : {e}"
            }
        })
    return JSONResponse({"documentResult": {"result": True, "status_code": "success",
                                            "description": "Document vectors deleted successfully"}}, status_code=200)


# CUDA 메모리 초기화 함수
def clear_cuda_memory():
    try:
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
            torch.cuda.ipc_collect()
            log("[clear_cuda_memory] CUDA memory cleared successfully.")
    except Exception as e:
        log(f"[clear_cuda_memory] Error clearing CUDA memory: {str(e)}")

# 전체 텍스트를 출력하는 함수
def process_raw_text(file_path: str, file_save_name: str, file_name: str):
    recursive = False
    parsed_text, chunk_metadata = extract_and_chunk_text_from_file(file_path, file_save_name, file_name, recursive)
    return parsed_text

# pdf 파일 페이지수 확인
def get_pdf_page_count(pdf_path: str) -> int:
    with pdfplumber.open(pdf_path) as pdf:
        return len(pdf.pages)

# 파일 형변환 to pdf
def convert_file_to_pdf(output_dir, target_file):
    subprocess.run(['soffice', '--headless', '--convert-to', 'pdf', '--outdir', output_dir, target_file])


def add_documents_milvus(file_path, file_save_name, file_name, splitter, start=0, end=-1, use_marker=True, category_pk=None,
        ocr_fallback=True, pdf_file_name='', category=''):
    if start > end:
        range = ""
    else:
        range = f"{start}~{end}"

    try:
        chunk_metadata = extract_and_chunk_text_from_file_range(
            file_path=file_path, file_save_name=file_save_name, file_name=file_name,
            splitter=splitter, start=start, end=end, use_marker=use_marker,
            category_pk=category_pk, ocr_fallback=ocr_fallback, pdf_file_name=pdf_file_name
        )
        if not chunk_metadata[1]:
            raise HTTPException(status_code=400, detail={
                "error": {
                    "description": "No content could be extracted from document"
                }
            })
        mark_documents = convert_to_documents(chunk_metadata, category)
        milvus_db.add_documents(mark_documents)
        log(f"[add_documents_milvus] {file_name} {range} Add document to milvus Done")
    except Exception as e:
        log(f"[add_documents_milvus] {file_name} {range} Error : {e}")

def extract_and_chunk_text_from_file_range(
        file_path, file_save_name, file_name, splitter, start=0, end=-1, use_marker=True, category_pk=None,
        ocr_fallback=True, pdf_file_name=''):

    all_chunks = []
    chunk_metadata = []
    parsed_text = ''
    def process_text_by_page(file_path, file_save_name, file_name, text_by_page_dict, all_chunks, chunk_metadata,
                                   splitter, category_pk):
        nonlocal parsed_text  # 외부 함수의 변수를 사용하기 위해 nonlocal 선언
        for page_number, page_data in text_by_page_dict.items():
            # 데이터가 list일때 : chunk를 만들어서 온 경우
            if isinstance(page_data, list):
                chunks = [attach_file_info(file_name, page_number, chunk) for chunk in page_data]
            else: # 데이터가 text인 경우 : chunk를 만든다.
                cleaned_text = get_cleaned_text(page_data)
                parsed_text += cleaned_text
                chunks = chunk_text(file_path, file_save_name, file_name, page_number - 1, cleaned_text,
                                    text_splitter=splitter)

            for chunk_number, chunk in enumerate(chunks):
                all_chunks.append(chunk)
                chunk_metadata.append({
                    "file_path": file_path,
                    "file_save_name": file_save_name,
                    "file_name": file_name,
                    "page_number": page_number,
                    "chunk_number": chunk_number + 1,
                    "text": chunk,
                    "category_pk": category_pk
                })

    if start > end:
        # 페이지 범위 유효하지 않음(PDF 파일 외)
        text_by_page_dict = extract_text_from_file(file_path, file_save_name, use_marker, ocr_fallback)
    else:
        # 페이지 범위 사용할때 (PDF)
        if use_marker:
            text_by_page_dict = preprocess_pdf_text_by_page_range(file_path, pdf_file_name, start, end, ocr_fallback)
        else:
            text_by_page_dict = extract_text_from_pdf(file_path, pdf_file_name, ocr_fallback)

    process_text_by_page(file_path, file_save_name, file_name, text_by_page_dict, all_chunks, chunk_metadata,
                                   splitter, category_pk)

    log(f"{file_name} 텍스트 추출 완료 {start}~{end}")
    return parsed_text, chunk_metadata


def preprocess_pdf_text_by_page_range(file_path, pdf_file_name, start, end, ocr_fallback=True):
    log(f"{pdf_file_name} ({start}-{end}) 전처리 처리 시작")
    pdf_full_path = os.path.join(file_path, pdf_file_name)
    # 해당 구간의 이미지&테이블 존재 여부 or 텍스트 추출
    analyzed_pages = analyze_pdf_range_structure(pdf_full_path, start, end, ocr_fallback)
    tmp_file_name_list = []

    # 메모리를 사용할지 디스크를 사용할지 여부
    use_memory = check_acl_permission(file_path)
    if use_memory:
        pdf_buffer = BytesIO()
    else:
        pdf_buffer = None

    # 일반 Marker를 사용할 페이지와 OCR Marker를 사용할 페이지 분리
    marker_target_pages = []
    ocr_target_pages = []
    for page_number, text in analyzed_pages.items():
        if text == 'has_image':
            ocr_target_pages.append(page_number)
        elif text == 'has_table':
            marker_target_pages.append(page_number)

    # 일반 marker 실행
    if len(marker_target_pages) > 0:
        tmp_file_name = make_tmp_pdf_file_select_pages(pdf_full_path, marker_target_pages, use_memory,
                                                              pdf_buffer)
        log(f"MarkerPdf용 파일 생성 Page {marker_target_pages} : {tmp_file_name}")
        results = extract_text_from_pdf_by_marker_pages(tmp_file_name, use_ocr=False)
        for r_idx, text in enumerate(results):
            page_number = marker_target_pages[r_idx]
            # 텍스트 검증
            if ocr_fallback and not is_valid_text(text):
                # 인코딩 깨진 경우, ocr 대상 페이지로 추가
                log(f"Page {page_number} 텍스트 깨짐 감지. OCR 재처리 대상")
                ocr_target_pages.append(page_number)
            else:
                analyzed_pages[page_number] = text
        tmp_file_name_list.append(tmp_file_name)

    # OCR marker 실행
    if len(ocr_target_pages) > 0:
        tmp_file_name = make_tmp_pdf_file_select_pages(pdf_full_path, ocr_target_pages, use_memory,
                                                              pdf_buffer)
        log(f"MarkerPdf(OCR)용 파일 생성 Page {ocr_target_pages} : {tmp_file_name}")
        results = extract_text_from_pdf_by_marker_pages(tmp_file_name, use_ocr=True)
        for r_idx, text in enumerate(results):
            page_number = ocr_target_pages[r_idx]
            analyzed_pages[page_number] = text
        tmp_file_name_list.append(tmp_file_name)

    if len(tmp_file_name_list)>0:
        for tmp_file_name in tmp_file_name_list:
            log(f"임시 파일 삭제 : {tmp_file_name}")
            os.remove(tmp_file_name)

    # 버퍼를 사용했다면, 닫아줌
    if use_memory and pdf_buffer:
        pdf_buffer.close()

    log(f"{pdf_file_name} ({start}-{end}) 전처리 처리 완료")
    return analyzed_pages

# 요청 바디를 위한 데이터 모델 정의
class DocumentRequest(BaseModel):
    kgptFileSaveName: str
    kgptFileName: str
    category: str = ""
    useMarker: bool = True
    categoryPk: str = ""
    splitterName: str = ""
    ocrFallback: bool = True # 인코딩 깨짐 확인후 OCR처리 여부


class MultipleMetadataRequest(BaseModel):
    prompt: str
    categoryList: List[str] = []
    kgptFileSaveNameList: List[str] = []
    topK: int = chunk_k
    useSearch: str = []
    categoryPk: List[str] = []

# 텍스트 임베딩을 위한 요청 모델
class TextEmbeddingRequest(BaseModel):
    content: str
    title: str


class MetricsInput(BaseModel):
    keyword1: str
    keyword2: str
    answer1: str
    answer2: str


class DeleteSessionRequest(BaseModel):
    sessionId: str


# 청킹된 데이터를 하나의 데이터로 만들기 위한 요청 모델(카테고리 내의 특정 문서)
class FullTextRequest(BaseModel):
    kgptFileSaveName: str
    categoryPk: str


# 청킹된 데이터를 하나의 데이터로 만들기 위한 요청 모델(카테고리 내의 5개 문서)
class FullTextListRequest(BaseModel):
    category: str


# 텍스트 문서를 처리하는 함수
def process_text_content(content: str, title: str):
    chunks = []
    chunk_metadata = []
    words = content.split()
    step = chunk_size - chunk_overlap
    text_chunks = [
        ' '.join(words[i:i + chunk_size])
        for i in range(0, len(words), step)
    ]
    for chunk_number, chunk in enumerate(text_chunks):
        chunks.append(chunk)
        chunk_metadata.append({
            "file_path": "text_contents",
            "file_save_name": f"{title}.txt",
            "file_name": title,
            "page_number": 1,
            "chunk_number": chunk_number + 1,
            "text": chunk
        })
    return chunks, chunk_metadata

@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request, exc: RequestValidationError):
    return JSONResponse(
        status_code=400,
        content={
            "detail": {
                "error": {
                    "description": "Required field missing: " + exc.errors()[0]['loc'][-1]
                }
            }
        }
    )

# 파싱 전체 결과 출력 엔드포인트
@app.post("/api-kgpt/grpc/process-text")
async def process_raw_text_endpoint(request: DocumentRequest):
    """
    텍스트 문서를 처리하고 결과를 반환하는 API
    """
    kgpt_file_save_name = request.kgptFileSaveName
    kgpt_file_name = request.kgptFileName
    category = request.category
    log("[process-text] kgpt_file_save_name : ", kgpt_file_save_name)
    log("[process-text] kgpt_file_name : ", kgpt_file_name)
    log("[process-text] category : ", category)

    # 파일 경로 생성
    file_path = os.path.join(category, kgpt_file_save_name)
    log("[process-text] pdf_path : ", file_path)
    log("[process-text] os.path.exists(pdf_path)", os.path.exists(file_path))

    # 파일 존재 여부 확인
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail={
            "documentResult": {
                "result": False,
                "status_code": "error",
                "description": f"File not found: {file_path}"
            }
        })

    try:
        # 텍스트 처리 함수 호출
        parsed_text = process_raw_text(os.path.dirname(file_path), kgpt_file_save_name, kgpt_file_name)
        log('parsed_text', parsed_text)
        # 결과 반환
        return JSONResponse(content={
            "documentResult": {
                "result": True,
                "status_code": "success",
                "description": "Document processed successfully",
                "parsedText": parsed_text
            }
        })

    except Exception as e:
        log(f"[process_raw_text_endpoint] Error processing raw text: {str(e)}")
        raise HTTPException(status_code=500, detail={
            "documentResult": {
                "result": False,
                "status_code": "error",
                "description": f"Failed to process raw text: {str(e)}"
            }
        })


@app.post("/api-kgpt/grpc/text-embedder")
async def text_embedder(request: TextEmbeddingRequest):
    if milvus_db is None:
        log("[load_or_create_embeddings] Milvus DB instance not found")
        raise HTTPException(status_code=404, detail="Milvus DB instance not found")

    try:
        content = request.content
        title = request.title
        log(f"[text_embedder] title: {title}")
        # 텍스트 처리 및 청킹
        chunk_metadata = process_text_content(content, title)
        mark_documents = convert_to_documents(chunk_metadata, "onlyText")
        milvus_db.add_documents(mark_documents)

        log(f"[text_embedder] success")
        return {
            "documentResult": {
                "result": True,
                "status_code": "success",
                "description": "Text content embedded successfully"
            }
        }
    except Exception as e:
        log(f"[text_embedder] Error embedding text content: {str(e)}")
        raise HTTPException(status_code=500, detail={
            "documentResult": {
                "result": False,
                "status_code": "error",
                "description": f"Failed to embed text content: {str(e)}"
            }
        })


# FastAPI 엔드포인트 정의(임베딩)
@app.post("/api-kgpt/grpc/document-embedder")
async def document_embedder(request: DocumentRequest):
    # 실행 시간 측정 시작
    start_time = time.time()
    log(f"[document_embedder_time_check Start] ######### time: {time.strftime("%H:%M:%S", time.localtime(start_time))} #########")

    kgpt_file_save_name = request.kgptFileSaveName
    kgpt_file_name = request.kgptFileName
    category = request.category
    use_marker = request.useMarker
    file_dir = request.categoryPk
    splitter_name = request.splitterName
    ocr_fallback = request.ocrFallback

    log("[document_embedder] kgpt_file_save_name ", kgpt_file_save_name)
    log("[document_embedder] kgpt_file_name ", kgpt_file_name)
    log("[document_embedder] category ", category)
    log("[document_embedder] use_marker ", use_marker)
    log("[document_embedder] category_pk ", file_dir)
    log("[document_embedder] splitter_name ", splitter_name)
    log("[document_embedder] ocr_fallback ", ocr_fallback)

    if milvus_db is None:
        log("[document_embedder] Milvus DB instance not found")
        raise HTTPException(status_code=404, detail="Milvus DB instance not found")

    # 파일 경로 설정 및 존재 여부 확인
    full_file_path = os.path.join(file_dir, kgpt_file_save_name)
    file_exist = os.path.exists(full_file_path)
    log("[document_embedder] file_path ", full_file_path)
    log("[document_embedder] os.path.exists(file_path)", file_exist)

    # 파일 존재 여부 확인
    if not file_exist:
        log(f"[document_embedder] File not found")
        raise HTTPException(status_code=404, detail={
            "error": {
                "description": "File not found"
            }
        })

    # 파일 사이즈 확인
    file_size = os.path.getsize(full_file_path)
    if file_size == 0:
        log(f"[document_embedder] Empty file")
        raise HTTPException(status_code=400, detail={
            "error": {
                "description": "Empty file"
            }
        })

    text_splitter = get_splitter(recursive=False, splitter_name=splitter_name)
    log(f"[document_embedder] 사용할 splitter : {type(text_splitter).__name__}")

    # 파일 경로와 파일명을 load_or_create_embeddings 함수로 전달
    try:
        if kgpt_file_save_name.lower().endswith(pdf_convert_file_ext + (".pdf",)):

            pdf_file_name = kgpt_file_save_name
            # pdf가 아닌 경우, pdf로 convert
            if not kgpt_file_save_name.lower().endswith('.pdf'):
                convert_file_to_pdf(output_dir=file_dir, target_file=full_file_path)
                pdf_file_name = os.path.splitext(kgpt_file_save_name)[0] + ".pdf"
                full_file_path = os.path.join(file_dir, pdf_file_name)

            pdf_file_length = get_pdf_page_count(full_file_path)

            if not use_marker:
                await handle_limited_thread(
                    add_documents_milvus,
                    file_path=file_dir, file_save_name=kgpt_file_save_name, file_name=kgpt_file_name,
                    splitter = text_splitter, start=0, end=pdf_file_length - 1, use_marker=use_marker,
                    category_pk=file_dir, ocr_fallback=ocr_fallback, pdf_file_name=pdf_file_name, category=category
                )
            else:
                # marker-pdf 사용 (병렬처리)
                # 각 스레드가 처리할 index 범위 (최소 10페이지 처리)
                range_length = max(10, math.ceil(pdf_file_length / max_thread_cnt))
                page_ranges = [
                    (i, min(i + range_length - 1, pdf_file_length - 1))
                    for i in range(0, pdf_file_length, range_length)
                ]

                # 각 스레드는 작업후 스레드 반납
                tasks = []
                for start, end in page_ranges:
                    task = asyncio.create_task(handle_limited_thread(
                        add_documents_milvus,
                        file_path=file_dir, file_save_name=kgpt_file_save_name, file_name=kgpt_file_name,
                        splitter=text_splitter, start=start, end=end, use_marker=use_marker,
                        category_pk=file_dir, ocr_fallback=ocr_fallback, pdf_file_name=pdf_file_name,
                        category=category
                    ))
                    tasks.append(task)

                task_results = await asyncio.gather(*tasks)
                log(f"[document_embedder] 모든 스레드 - 태스크 작업 완료 result length : {len(task_results)}")

                # convert 했을 경우, 임시 pdf파일 삭제
                if not kgpt_file_save_name.lower().endswith('.pdf'):
                    try:
                        os.remove(full_file_path)
                        log(f"[document_embedder] Deleted 임시 PDF file: {pdf_file_name}")
                    except OSError as e:
                        log(f"[document_embedder] Error deleting 임시 PDF file: {pdf_file_name}, {e}")
        else:
            await handle_limited_thread(
                add_documents_milvus,
                file_path=file_dir, file_save_name=kgpt_file_save_name, file_name=kgpt_file_name,
                splitter=text_splitter, use_marker=use_marker, category_pk=file_dir, category=category
            )

        log(f"[document_embedder] Document '{kgpt_file_name}' successfully embedded in Milvus.")

    except HTTPException as he:
        # HTTPException의 detail을 그대로 전달
        log(f"[document_embedder] HTTPException embedding document: {str(he)}")
        raise he
        
    except Exception as e:
        log(f"[document_embedder] Error embedding document: {str(e)}")
        raise HTTPException(status_code=500, detail={
            "error": {
                "description": str(e)
            }
        })

    # cuda 메모리 초기화
    clear_cuda_memory()

    # 종료
    end_time = time.time()
    log(f"[document_embedder_time_check End] ######### time: {time.strftime("%H:%M:%S", time.localtime(end_time))} #########")
    elapsed_time = end_time - start_time
    log(f"[document_embedder_time_check Total time] ######### Total: {elapsed_time:.2f} seconds #########")

    return {
        "documentResult": {
            "result": True,
            "status_code": "success",
            "description": "Document embedded successfully"
        }
    }


# FastAPI 엔드포인트 정의, pdf를 마커 없이임베딩
@app.post("/api-kgpt/grpc/document-embedder-n-marker")
async def document_embedder_not_marker(request: DocumentRequest):
    kgpt_file_save_name = request.kgptFileSaveName
    kgpt_file_name = request.kgptFileName
    category = request.category

    log("[document_embedder] kgpt_file_save_name ", kgpt_file_save_name)
    log("[document_embedder] kgpt_file_name ", kgpt_file_name)
    log("[document_embedder] category ", category)

    # 파일 경로 설정 및 존재 여부 확인
    pdf_path = os.path.join(category, kgpt_file_save_name)
    log("[document_embedder] pdf_path ", pdf_path)
    log("[document_embedder] os.path.exists(pdf_path)", os.path.exists(pdf_path))
    if not os.path.exists(pdf_path):
        raise HTTPException(status_code=404, detail={
            "documentResult": {
                "result": False,
                "status_code": "error",
                "description": "File not found"
            }
        })

    # 파일 경로와 파일명을 load_or_create_embeddings 함수로 전달
    try:
        load_or_create_embeddings(
            file_path=os.path.dirname(pdf_path),  # 파일의 디렉토리 경로
            file_save_name=kgpt_file_save_name,
            file_name=kgpt_file_name,
            category=category,
            recursive=False,
            use_marker=False
        )
        log(f"[document_embedder] Document '{kgpt_file_name}' successfully embedded in Milvus.")


    except Exception as e:
        log(f"[document_embedder] Error embedding document: {str(e)}")
        raise HTTPException(status_code=500, detail={
            "documentResult": {
                "result": False,
                "status_code": "error",
                "description": f"Failed to embed document: {str(e)}"
            }
        })

    # cuda 메모리 초기화
    clear_cuda_memory()

    return {
        "documentResult": {
            "result": True,
            "status_code": "success",
            "description": "Document embedded successfully"
        }
    }


# FastAPI 엔드포인트 정의(임베딩) - splitter가 recursive
@app.post("/api-kgpt/grpc/document-embedder-re")
async def document_embedder_recursive(request: DocumentRequest):
    kgpt_file_save_name = request.kgptFileSaveName
    kgpt_file_name = request.kgptFileName
    category = request.category

    log("[document_embedder] kgpt_file_save_name ", kgpt_file_save_name)
    log("[document_embedder] kgpt_file_name ", kgpt_file_name)
    log("[document_embedder] category ", category)

    # 파일 경로 설정 및 존재 여부 확인
    pdf_path = os.path.join(category, kgpt_file_save_name)
    log("[document_embedder] pdf_path ", pdf_path)
    log("[document_embedder] os.path.exists(pdf_path)", os.path.exists(pdf_path))
    if not os.path.exists(pdf_path):
        raise HTTPException(status_code=404, detail={
            "documentResult": {
                "result": False,
                "status_code": "error",
                "description": "File not found"
            }
        })

    # 파일 경로와 파일명을 load_or_create_embeddings 함수로 전달
    try:
        load_or_create_embeddings(
            file_path=os.path.dirname(pdf_path),  # 파일의 디렉토리 경로
            file_save_name=kgpt_file_save_name,
            file_name=kgpt_file_name,
            category=category,
            recursive=True,
            use_marker=True
        )
        log(f"[document_embedder] Document '{kgpt_file_name}' successfully embedded in Milvus.")


    except Exception as e:
        log(f"[document_embedder] Error embedding document: {str(e)}")
        raise HTTPException(status_code=500, detail={
            "documentResult": {
                "result": False,
                "status_code": "error",
                "description": f"Failed to embed document: {str(e)}"
            }
        })

    # cuda 메모리 초기화
    clear_cuda_memory()

    return {
        "documentResult": {
            "result": True,
            "status_code": "success",
            "description": "Document embedded successfully"
        }
    }


# FastAPI 엔드포인트 정의
@app.post('/api-kgpt/grpc/document-delete')
async def document_delete(request_data: DocumentRequest):
    kgpt_file_save_name = request_data.kgptFileSaveName
    kgpt_file_name = request_data.kgptFileName
    category = request_data.category
    category_pk = request_data.categoryPk

    log("[document_delete] kgpt_file_save_name ", kgpt_file_save_name)
    log("[document_delete] kgpt_file_name ", kgpt_file_name)
    log("[document_delete] category ", category)
    log("[document_delete] category_pk ", category_pk)

    response = await handle_limited_thread(delete_document_from_milvus, file_save_name=kgpt_file_save_name)
    return response


@app.post('/api-kgpt/grpc/delete_test_session')
async def delete_test_session(request_data: DeleteSessionRequest):
    session_id = request_data.sessionId
    log("[delete_test_session] session_id ", session_id)
    return delete_document_from_milvus_by_session_id(session_id)


# multiple_metadata FastAPI 엔드포인트 정의
@app.post("/api-kgpt/grpc/multiple_metadata")
async def multiple_metadata(request: MultipleMetadataRequest):
    # 실행 시간 측정 시작
    start_time = time.time()
    log(f"[document_retriever_time_check Start] ######### time: {time.strftime("%H:%M:%S", time.localtime(start_time))} #########")

    prompt = request.prompt
    category_list = request.categoryList
    file_save_name_list = request.kgptFileSaveNameList
    top_k = request.topK
    use_keyword = request.useSearch
    category_pk = request.categoryPk

    # 빈 프롬프트 체크
    if not prompt:
        raise HTTPException(status_code=404, detail="Prompt is empty")
    log("[multiple_metadata] prompt ", prompt)
    log("[multiple_metadata] categoryList ", category_list)
    log("[multiple_metadata] fileSaveNameList ", file_save_name_list)
    log("[multiple_metadata] top_k ", top_k)
    log("[multiple_metadata] use_keyword ", use_keyword)
    log("[multiple_metadata] category_pk ", category_pk)

    # Milvus 인스턴스 확인
    if milvus_db is None:
        log("[multiple_metadata] Milvus DB instance not found")
        raise HTTPException(status_code=404, detail="Milvus DB instance not found")

    # Retriever 생성 및 Chain 생성
    try:
        # use_keyword가 "search"일 경우 키워드 기반 검색, 그렇지 않을 경우 기본 검색
        if use_keyword == "search":
            log("[multiple_metadata] Using keyword-based retriever")
            retriever = await handle_limited_thread(keyword_search,
                                category_list=category_list, file_save_name_list=file_save_name_list, top_k=top_k)
        else:
            log("[multiple_metadata] Using default retriever")
            retriever = await handle_limited_thread(get_reranker,
                                    file_save_name_list=file_save_name_list, top_n=top_k, category_pk=category_pk)


    except Exception as e:
        log(f"[multiple_metadata] Error occurred while initializing the retriever: {e}")
        raise HTTPException(status_code=500, detail={
            "error": {
                "description": "Failed to initialize retriever."
            }
        })

    try:
        # 프롬프트에 대해 chain 실행하여 검색된 문서 가져오기
        retrieved_docs = retriever.invoke(prompt)
        log("[generate_question_from_chunks] chunk_k ", len(retrieved_docs))

        # 문서 필터링: category_list에 포함된 파일명과 일치하는 것만 필터링
        # relevant_chunks = [
        #    doc for doc in retrieved_docs
        #    if doc.metadata.get('file_path') in category_list]
        # log("최종 Chroma_result 개수 : ", len(relevant_chunks))

        # 사용할 문서 선택 (필터링된 청크가 있으면 사용하고, 없으면 원본 사용)
        # response_chunks = relevant_chunks if len(relevant_chunks) > 0 else retrieved_docs

        # 응답 데이터 포맷팅
        response = convert_to_serializable(retrieved_docs)
        if not response:
            log("[multiple_metadata] No relevant documents found after filtering")
        log("[RESPONSE] >>>> >>>> >>>> >>> \n", json.dumps(response, indent=2, ensure_ascii=False))
        log("=" * 124)
        log("=" * 124)

        # 종료
        end_time = time.time()
        log(f"[document_retriever_time_check End] ######### time: {time.strftime("%H:%M:%S", time.localtime(end_time))} #########")
        elapsed_time = end_time - start_time
        log(f"[document_retriever_time_check Total time] ######### Total: {elapsed_time:.2} seconds #########")

        return JSONResponse(content=response)

    except Exception as e:
        log(f"[multiple_metadata] Error retrieving relevant chunks: {e}")
        raise HTTPException(status_code=500, detail="Failed to retrieve relevant chunks")


# 특정 문서 하나의 청크 데이터를 하나의 텍스트로 가져오는 FastAPI 엔드포인트 정의
@app.post("/api-kgpt/grpc/full_text_metadata")
async def full_text_metadata(request: FullTextRequest):
    kgpt_file_save_name = request.kgptFileSaveName
    categoryPk = request.categoryPk

    log("[full_text_metadata] kgptFileSaveName ", kgpt_file_save_name)
    log("[full_text_metadata] categoryPk ", categoryPk)

    # Milvus 인스턴스 확인
    if milvus_db is None:
        log("[full_text_metadata] Milvus DB instance not found")
        raise HTTPException(status_code=404, detail="Milvus DB instance not found")

    try:
        response = await handle_limited_thread(get_full_chunk_text, categoryPk=categoryPk, file_save_name=kgpt_file_save_name)
        return JSONResponse(content=response)

    except Exception as e:
        log(f"[full_text_metadata] Error retrieving relevant chunks: {e}")
        raise HTTPException(status_code=500, detail={
            "error": {
                "description": "Failed to find file"
            }
        })


# 카테고리에서 문서 5개를 통으로 리스트에 가져오는 FastAPI 엔드포인트 정의
@app.post("/api-kgpt/grpc/full_text_multiple_metadata")
async def full_text_multiple_metadata(request: FullTextListRequest):
    category = request.category
    log("[full_text_multiple_metadata] category ", category)

    if milvus_db is None:
        log("[full_text_multiple_metadata] Milvus DB instance not found")
        raise HTTPException(status_code=404, detail="Milvus DB instance not found")

    try:
        response = await handle_limited_thread(get_full_chunk_text_list, category=category)
        return JSONResponse(content=response)

    except Exception as e:
        log(f"[full_text_multiple_metadata] Error retrieving relevant chunks: {e}")
        raise HTTPException(status_code=500, detail="Failed to find file")

# 컬렉션삭제후 새로 컬렉션 생성
@app.post("/api-kgpt/grpc/reset_milvus")
async def reset_milvus_collection():
    global milvus_db
    try:
        log(f"[reset_milvus] Collection : {milvus_db.get_collection_name()}")
        milvus_db = reset_milvus(milvus_db)
    except Exception as e:
        log(f"[reset_milvus] Error : {e}")
        raise HTTPException(status_code=500, detail="Failed to reset milvus db")

    return JSONResponse({"documentResult": {"result": True, "status_code": "success",
                                            "description": "Reset milvus db successfully"}}, status_code=200)

# 해당 db에 컬렉션명 조회
@app.post("/api-kgpt/grpc/show_milvus_collection")
async def show_milvus_collection():
    global milvus_db
    try:
        response = milvus_db.get_collections()
        log(f"[show_milvus_collection] Collection : {response}")
    except Exception as e:
        log(f"[reset_milvus] Error : {e}")
        raise HTTPException(status_code=500, detail="Failed to show milvus collections")
    return JSONResponse(content=response)

if __name__ == "__main__":
    nest_asyncio.apply()
    uvicorn.run(app, host="0.0.0.0", port=5000, reload=False)