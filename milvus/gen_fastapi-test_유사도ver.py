import os
import time
import pandas as pd
import fitz
import numpy as np
import logging
from logging.handlers import RotatingFileHandler
import inspect
import json
import subprocess
from bs4 import BeautifulSoup
from typing import List, Dict, Union
from logging.handlers import RotatingFileHandler
from langchain_huggingface import HuggingFaceEmbeddings
from langchain.retrievers import ContextualCompressionRetriever
from langchain.retrievers.document_compressors import CrossEncoderReranker
from langchain_community.retrievers import BM25Retriever
from langchain_community.retrievers import TFIDFRetriever
from langchain_community.cross_encoders import HuggingFaceCrossEncoder
from langchain.text_splitter import MarkdownHeaderTextSplitter
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_experimental.text_splitter import SemanticChunker
from langchain_chroma import Chroma
from langchain.schema import Document
from fastapi import FastAPI, HTTPException
from fastapi.responses import JSONResponse
from langchain_community.document_loaders import UnstructuredExcelLoader
from langchain.retrievers import EnsembleRetriever
from langchain_community.document_loaders.csv_loader import CSVLoader
from langchain_community.document_loaders import UnstructuredHTMLLoader
from pydantic import BaseModel
import nest_asyncio
import uvicorn
from HWPLoader import HWPLoader
import torch
import warnings
import re
from sklearn.metrics import precision_score, recall_score, accuracy_score, f1_score
from nltk.translate.meteor_score import single_meteor_score
import nltk
from openpyxl import load_workbook

# 마커 새로운버전
from marker.converters.pdf import PdfConverter
from marker.models import create_model_dict
from marker.output import text_from_rendered
from marker.config.parser import ConfigParser
from sympy.integrals.meijerint_doc import category

warnings.filterwarnings("ignore", message=".*could benefit from vacuuming your database.*")

nltk.download('wordnet')
nltk.download('punkt')

# UNstructed HTML로더 사용 시 필수 nltk
nltk.download('punkt_tab') 
nltk.download('averaged_perceptron_tagger')
nltk.download('averaged_perceptron_tagger_eng')


app = FastAPI()

# 전역 변수 정의 및 초기화 함수
chroma_db = None

# model_name = "BAAI/bge-m3"  # 로컬 테스트용
model_name = "/home/app/bge-m3"

# 설치가 필요한 cross-encoder 모델
# reranker_name = "BAAI/bge-reranker-v2-m3" # 로컬 테스트용
reranker_name = "/home/app/bge-reranker-v2-m3"

model_kwargs = {"device": "cuda"}
encode_kwargs = {"normalize_embeddings": True}

chunk_size = 500  # 텍스트 청크 사이즈
chunk_overlap = 20  # 청크 중복
chunk_k = 20  # 검색 결과 수
max_ocr_page_count = 5000
MAX_THREADS = os.cpu_count()  # CPU 코어 개수에 따라 동적으로 MAX_THREADS 설정, 시스템의 CPU 코어 개수로 설정
# 청크 데이터 관련 전역 변수
all_chunks = None
chunk_metadata = None

# RAG Name
rag_name = "gen_rag"  # lower case
reg_version = ""
# 설정
data_dir = f"data_{rag_name}{reg_version}"  # PDF 파일 디렉토리

# 기본 설정 값
DB_PATH = "Chroma_Markdown"
COLLECTION_NAME = "my_db"

# Parsing 모델 로드하기
# model_list = load_all_models()

headers_to_split_on = [("####", "Page Of Content")]
markdown_splitter = MarkdownHeaderTextSplitter(headers_to_split_on=headers_to_split_on)
recursive_splitter = RecursiveCharacterTextSplitter()

# 로거 초기화
logger = logging.getLogger(__name__)

# 마커 새로운버전
converter = PdfConverter(
    artifact_dict=create_model_dict(),
)

# pdf로 변환할 파일 확장자
pdf_convert_file_ext = (".doc", ".docx", ".ppt", ".pptx", ".hwpx")

# Set up a rotating file handler and console handler
def initialize_logger():
    global logger
    logger.setLevel(logging.DEBUG)
    # 기존 핸들러 제거
    logger.handlers = []

    # File handler with rotating logs
    file_handler = RotatingFileHandler(
        filename='gen_rag.log',
        maxBytes=100 * 1024 * 1024,  # 100 MB
        backupCount=2,
        encoding='utf-8'
    )

    # Console handler
    console_handler = logging.StreamHandler()

    # Format for the log messages with detailed timestamp including milliseconds
    formatter = logging.Formatter(
        '[%(asctime)s.%(msecs)03d][%(levelname)s] %(message)s - %(custom_filename)s:%(custom_lineno)d',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    # Add handlers to the logger
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    # 루트 로거로의 전파 방지
    logger.propagate = False


initialize_logger()


# Logging function with added error handling
def log(*args, **kwargs):
    try:
        # 로그 메시지를 문자열로 변환
        log_message = ' '.join(map(str, args))

        # 호출자 정보 추출
        frame = inspect.currentframe().f_back
        file_name = os.path.basename(frame.f_code.co_filename)
        line_number = frame.f_lineno

        # 로그 메시지 기록 (extra 인자에 커스텀 키 사용)
        logger.debug(log_message, extra={'custom_filename': file_name, 'custom_lineno': line_number})

    except Exception as e:
        logger.error(f"[log] Error logging message: {e}")


# 임베딩 모델 로드하기
def load_embedding_model():
    log("[load Model Start]")
    # model = SentenceTransformer(model_name, device='cuda')
    model = HuggingFaceEmbeddings(
        model_name=model_name, model_kwargs=model_kwargs, encode_kwargs=encode_kwargs)
    log("[load Model End]")
    return model


# Chroma 데이터베이스 초기화 함수
def init_chroma():
    global chroma_db

    if chroma_db is None:
        chroma_db = Chroma(persist_directory=DB_PATH, embedding_function=embedding_model,
                           collection_name=COLLECTION_NAME)


# 파일 내용 가공
def get_cleaned_text(text):
    cleaned_text = text
    return cleaned_text


# embedding 모델 로드하기
embedding_model = load_embedding_model()

# 시맨틱 청킹 수행
splitter_semantic = SemanticChunker(embedding_model)


# 파일 정보를 page_content에 붙여, 양질의 검색 수행
def attach_file_info(directory, file_save_name, file_name, page_number, text):
    if isinstance(text, Document):
        text = text.page_content
    return f"####Record ID: {file_name}, page Number: {page_number}\n\n" + text


# 한글(HWP) 파일에서 텍스트 추출
def extract_text_from_hwp(file_path: str) -> str:
    hwp_loader = HWPLoader(file_path)
    document = next(hwp_loader.lazy_load())
    return document.page_content


# PDF에서 텍스트 추출 함수
def extract_text_from_pdf(pdf_path):
    doc = fitz.open(pdf_path)
    text_by_page = []
    for page_num in range(doc.page_count):
        page = doc.load_page(page_num)
        text = page.get_text("text")
        text_by_page.append(text)
    return text_by_page


# 텍스트 스플리터(recursive에 따라 스플리터 변경) 수행
def chunk_text(directory, file_save_name, file_name, page_num, text, recursive):
    filename, ext = os.path.splitext(os.path.basename(file_save_name))

    # recursive가 참이면 recursive splitter를 사용하고 그렇지 않다면 markdown splitter를 사용
    text_splitter = recursive_splitter if recursive else markdown_splitter

    chunks = text_splitter.split_text(text)

    chunked_texts = []

    # 청킹이 수행된 각 텍스트의 인덱스에 위에 설정한 파일명, 페이지 번호, 청킹 결과
    for chunk in chunks:
        chunked_texts.append(attach_file_info(directory, file_save_name, filename, page_num + 1, chunk))
    return chunked_texts

# 텍스트를 청크 사이즈 기준으로 자름
def split_text_by_chunk_size(text, chunk_size=chunk_size, chunk_overlap=chunk_overlap):
    step = chunk_size - chunk_overlap
    words = text.split()
    all_chunks = [
        ' '.join(words[i:i+chunk_size])
        for i in range(0, len(words), step)
    ]
    return all_chunks

# 재귀적으로 html 테이블 파싱
def parse_html_table(tag, level=0):
    rows = []
    for row in tag.find_all('tr', recursive=False):
        cells = []
        for cell in row.find_all(['th', 'td'], recursive=False):
            # 하위 테이블이 있을 경우 재귀적으로 호출
            nested_table = cell.find('table')
            if nested_table:
                parsed_tables = parse_html_table(nested_table, level + 1)
                cells.append(parsed_tables)
            else:
                cells.append(cell.get_text(strip=True))
        rows.append(cells)
    table_rows = [dict(zip(rows[0], row)) for row in rows[1:]]
    return json.dumps(table_rows, ensure_ascii=False)

# html파일에서 chunks를 리턴
def extract_chunks_and_table_from_html(html_path):
    with open(html_path, "r", encoding="utf-8") as file:
        soup = BeautifulSoup(file, "html.parser")

    html_text = ""
    tmp_text = ""
    html_chunks = []

    for tag in soup.find_all(['span', 'p', 'table']):
        if tag.name in ['span', 'p']:
            tmp_text = tag.get_text(strip=True)
            html_text += tmp_text + "\n"

        elif tag.name == 'table':
            # 기존 text는 청크로 저장
            chunks = split_text_by_chunk_size(html_text)
            # print("\n1 chunks", chunks)
            html_chunks += chunks
            # print("\n2 chunks", chunks)
            html_text = ""
            
            # table은 json으로 만든후 청크로 저장
            parsed_json_text = parse_html_table(tag)
            # 청크 제일 앞에 이전 문자열(표 제목) 추가
            html_chunks.append(tmp_text+"\n"+parsed_json_text)

    if html_text != "":
        chunks = split_text_by_chunk_size(html_text)
        html_chunks += chunks

    return html_chunks


# 파싱 + 청킹 수행 함수
def extract_and_chunk_text_from_file(file_path, file_save_name, file_name, recursive, use_marker=True, category_pk=None):
    all_chunks = []
    chunk_metadata = []
    parsed_text = ''

    def process_text_by_page(file_path, file_save_name, file_name, text_by_page, all_chunks, chunk_metadata, recursive):
        nonlocal parsed_text  # 외부 함수의 변수를 사용하기 위해 nonlocal 선언
        for page_num, text in enumerate(text_by_page):
            cleaned_text = get_cleaned_text(text)
            parsed_text += cleaned_text
            chunks = chunk_text(file_path, file_save_name, file_name, page_num, cleaned_text, recursive)
            for chunk_number, chunk in enumerate(chunks):
                all_chunks.append(chunk)
                chunk_metadata.append({
                    "file_path": file_path,
                    "file_save_name": file_save_name,
                    "file_name": file_name,
                    "page_number": page_num + 1,
                    "chunk_number": chunk_number + 1,
                    "text": chunk,
                    "category_pk": category_pk
                })

    # pdf로 변환할 파일 확장자일 경우
    if file_save_name.lower().endswith(pdf_convert_file_ext):

        def convert_to_pdf(input_file, output_file):
            subprocess.run(['soffice', '--headless', '--convert-to', 'pdf', '--outdir', os.path.join(file_path),
                            os.path.join(file_path, file_save_name)])

        convert_to_pdf(f"{file_path}{file_save_name}", file_path)

        pdf_file_name = os.path.splitext(file_save_name)[0] + ".pdf"

        # pdf에서 텍스트 추출
        text_by_page = preprocess_pdf_text_by_page(file_path, pdf_file_name, pdf_file_name, use_marker)
        process_text_by_page(file_path, file_save_name, file_name, text_by_page, all_chunks, chunk_metadata, recursive)

        # PDF 파일 삭제
        try:
            os.remove(os.path.join(file_path, pdf_file_name))
            log(f"[extract_and_chunk_text_from_file] Deleted PDF file: {pdf_file_name}")
        except OSError as e:
            log(f"[extract_and_chunk_text_from_file] Error deleting file: {pdf_file_name}, {e}")

    # PDF
    elif file_save_name.lower().endswith(".pdf"):
        text_by_page = preprocess_pdf_text_by_page(file_path, file_save_name, file_name, use_marker)
        process_text_by_page(file_path, file_save_name, file_name, text_by_page, all_chunks, chunk_metadata, recursive)

    # 엑셀 파일(xlsx, xls) 처리
    elif file_save_name.lower().endswith((".xlsx", ".xls")):
        wb = load_workbook(filename=os.path.join(file_path, file_save_name), read_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.rows)
            header_index = 0
            
            for idx, row in enumerate(rows):
                if any(cell.value is not None for cell in row):
                    header_index = idx
                    break

            headers = [cell.value for cell in rows[header_index]]
            row_strings = []

            for row in rows[header_index + 1:]:
                if not any(cell.value for cell in row):
                    continue
                row_string = ""
                for header, cell in zip(headers, row):
                    value = str(cell.value) if cell.value is not None else ''
                    if value:
                        row_string += f"{header} : {value}\n"

                if row_string.strip():
                    row_strings.append(row_string)


            for cell_index, cell in enumerate(row_strings):
                all_chunks.append(cell)
                chunk_metadata.append({
                    "file_path": file_path,
                    "file_save_name": file_save_name,
                    "file_name": file_name,
                    "page_number": 1,
                    "chunk_number": cell_index + 1,
                    "text": cell,
                    "category_pk": category_pk
                })
        wb.close()

    # CSV 파일 처리
    elif file_save_name.lower().endswith(".csv"):
        loader = CSVLoader(file_path=os.path.join(file_path, file_save_name))
        docs = loader.load()
        parsed_text = "\n".join([doc.page_content for doc in docs])
        for doc_index, doc in enumerate(docs):
            all_chunks.append(doc.page_content)
            chunk_metadata.append({
                "file_path": file_path,
                "file_save_name": file_save_name,
                "file_name": file_name,
                "page_number": 1,
                "chunk_number": doc_index + 1,
                "text": doc.page_content,
                "category_pk": category_pk
            })

    elif file_save_name.lower().endswith('.hwp'):
        result_text = extract_text_from_hwp(os.path.join(file_path, file_save_name))
        parsed_text = result_text
        chunks = chunk_text(file_path, file_save_name, file_name, 0, result_text, recursive)
        for chunk_number, chunk in enumerate(chunks):
            all_chunks.append(chunk)
            chunk_metadata.append({
                "file_path": file_path,
                "file_save_name": file_save_name,
                "file_name": file_name,
                "page_number": 1,  # HWP 파일에는 페이지 정보가 없으므로 1로 설정
                "chunk_number": chunk_number + 1,
                "text": chunk,
                "category_pk": category_pk
            })

    elif file_save_name.lower().endswith(".htm") or file_save_name.lower().endswith(".html"):
        #loader = UnstructuredHTMLLoader(os.path.join(file_path, file_save_name))
        #docs = loader.load()
        #parsed_text = "\n".join([doc.page_content for doc in docs])
        #html_split = splitter_semantic.split_documents(docs)
        #for doc_index, doc in enumerate(html_split):
        #    all_chunks.append(doc.page_content)

        # html데이터 추출
        html_chunks = extract_chunks_and_table_from_html(os.path.join(file_path, file_save_name))
        # print("html_chunks", html_chunks)
        parsed_text = "\n".join(html_chunks)
        for chunk_number, chunk in enumerate(html_chunks):
            all_chunks.append(chunk)
            chunk_metadata.append({
                "file_path": file_path,
                "file_save_name": file_save_name,
                "file_name": file_name,
                "page_number": 1, # 페이지는 무조건 1
                "chunk_number": chunk_number + 1,
                "text": chunk,
                "category_pk": category_pk
            })

    # TXT
    elif file_save_name.lower().endswith(".txt"):
        with open(os.path.join(file_path, file_save_name), 'r', encoding='utf-8') as f:
            text = f.read()
        parsed_text += text
        all_chunks = split_text_by_chunk_size(text)
        chunk_metadata = [
            {
                "file_path": file_path,
                "file_save_name": file_save_name,
                "file_name": file_name,
                "page_number": 1,  # TXT 파일에는 페이지 개념이 없으므로 1로 설정
                "chunk_number": i + 1,
                "text": chunk,
                "category_pk": category_pk
            }
            for i, chunk in enumerate(all_chunks)
        ]

    return parsed_text, chunk_metadata


# PDF 파일의 전처리 여부를 받아, text return
def preprocess_pdf_text_by_page(file_path, file_save_name, file_name, use_marker):
    text_by_page = None
    if use_marker:
        fname = file_path + "/" + file_save_name
        batch_multiplier = 12
        lang = ['ko', 'en']
        rendered = converter(fname)
        text_by_page, _, _ = text_from_rendered(rendered)
    else:
        text_by_page = extract_text_from_pdf(os.path.join(file_path, file_save_name))
    return text_by_page


# 청크 데이터를 Document 객체로 변환하는 함수 추가
def convert_to_documents(chunked_data, category):
    documents = []

    for chunk in chunked_data[1]:  # chunked_data[1]은 메타데이터 리스트
        # 메타데이터 정보 설정
        metadata = {
            "file_path": chunk["file_path"],
            "file_save_name": chunk["file_save_name"],
            "file_name": chunk["file_name"],
            "page_number": chunk["page_number"],
            "chunk_number": chunk["chunk_number"],
            "category": category,
            "category_pk": chunk["category_pk"]
        }

        # Document 생성 (page_content는 text로 설정)
        document = Document(page_content=chunk["text"], metadata=metadata)
        documents.append(document)

    return documents


def load_or_create_embeddings(file_path, file_save_name, file_name, category, recursive, use_marker, category_pk, db_path=DB_PATH,
                              collection_name=COLLECTION_NAME):
    """
    캐싱된 Chroma DB를 로드하거나, 새로운 임베딩을 생성하여 저장합니다.
    """
    # global bm25_retriever

    # 파일 식별용 키 생성
    document_id = os.path.join(file_name)

    # Chroma DB 로드
    if os.path.exists(db_path):
        log("기존 Chroma DB 로드 중입니다.")
        chroma_db = Chroma(
            persist_directory=db_path,
            embedding_function=embedding_model,
            collection_name=collection_name
        )

        # 기존 DB에 문서가 있는지 확인 (캐싱 여부 판단)
        stored_metadata = chroma_db.get(include=['metadatas'])
        existing_files = [
            doc['file_name'] for doc in stored_metadata['metadatas']
            if 'file_name' in doc
        ]

        # 이미 캐싱된 파일이면, 로드 후 반환
    #        if document_id in existing_files:
    #            log(f"문서 '{file_name}'는 이미 캐싱되어 있습니다.")
    #            return chroma_db  # 기존 DB 반환

    else:
        log("새로운 Chroma DB 생성 중 입니다.")
        chroma_db = None

    # 새로운 파일을 캐싱해야 하는 경우에만 파싱 및 청킹 수행
    log(f"문서 '{file_name}'에 대해 새로 파싱 및 청킹을 수행합니다.")
    chunk_metadata = extract_and_chunk_text_from_file(file_path, file_save_name, file_name, recursive, use_marker, category_pk)

    mark_documents = convert_to_documents(chunk_metadata, category)

    # 새로운 문서를 임베딩하여 Chroma DB 생성 또는 업데이트
    log(f"문서 '{file_name}'에 대해 새로 임베딩 수행 중입니다.")
    chroma_db = Chroma.from_documents(
        mark_documents, embedding_model,
        persist_directory=db_path,
        collection_name=collection_name
    )

    # return chroma_db, bm25_retriever


def get_full_chunk_text_list(chroma_db, category):
    chroma_data = chroma_db.get(include=["documents", "metadatas"])
    stored_document = chroma_data["documents"]
    stored_metadata = chroma_data["metadatas"]

    # file_dict 생성
    file_dict = {}

    for idx, meta in enumerate(stored_metadata):
        file_save_name = meta['file_save_name']
        if meta["file_path"].startswith(category):
            if file_save_name not in file_dict.keys() and len(file_dict.keys()) < 5:
                file_dict[file_save_name] = []
                file_dict[file_save_name].append([stored_document[idx], meta])
            elif file_save_name in file_dict.keys():
                file_dict[file_save_name].append([stored_document[idx], meta])
    data_list = []

    for file_name_key in file_dict.keys():
        sorted_filtered_metadata = sorted(file_dict[file_name_key], key=lambda x: x[1]['page_number'], reverse=False)

        selected_data_dict = {
            'full_text': '',
            'file_path': sorted_filtered_metadata[0][1]['file_path'],
            'file_save_name': sorted_filtered_metadata[0][1]['file_save_name'],
        }

        sum_page_contents = [page_content[0] for page_content in sorted_filtered_metadata]
        selected_data_dict['full_text'] = '\n\n\n'.join(sum_page_contents)
        data_list.append(selected_data_dict)

    return data_list


def get_full_chunk_text(chroma_db, category, file_save_name):
    chroma_data = chroma_db.get(include=["documents", "metadatas"])
    stored_document = chroma_data["documents"]
    stored_metadata = chroma_data["metadatas"]

    filtered_metadata = [
        [stored_document[idx], meta]
        for idx, meta in enumerate(stored_metadata)
        if meta["file_path"].startswith(category) and meta["file_save_name"] == file_save_name
    ]

    sorted_filtered_metadata = sorted(filtered_metadata, key=lambda x: x[1]['page_number'], reverse=False)

    selected_data_dict = {
        'full_text': '',
        'file_path': sorted_filtered_metadata[0][1]['file_path'],
        'file_save_name': sorted_filtered_metadata[0][1]['file_save_name'],
        'file_name': sorted_filtered_metadata[0][1]['file_name'],
    }

    sum_page_contents = [page_content[0] for page_content in sorted_filtered_metadata]
    selected_data_dict['full_text'] = '\n\n\n'.join(sum_page_contents)
    log("[full_text_metadata] text \n\n", selected_data_dict['full_text'])
    return selected_data_dict


# Retriever 설정 함수 수정
def get_reranker(chroma_db, category_list=[], file_save_name_list=[], top_n=chunk_k, k=20, fetch_k=30, category_pk=None):
    stored_metadata = chroma_db.get(include=["metadatas"])["metadatas"]
    text = chroma_db.get()['documents']

    log(f"[get_reranker] category_pk : \n {category_pk}")

    # bm25용 데이터
    bm25_converted_data = []
    # chroma용 쿼리 필터
    filtered_paths, filtered_files = set(), set()

    for document, meta in zip(text, stored_metadata):
        
        # 파일 경로가 category_list에 포함되는지 확인
        # if any(meta["file_path"].startswith(category) for category in category_list):
        #     filtered_paths.add(meta["file_path"])
        #     meta['text'] = document
        #     bm25_converted_data.extend(convert_to_documents([[], [meta]], meta['category']))

        # 폴더 pk 기준 조회
        if meta.get("category_pk") in category_pk:
            meta['text'] = document
            bm25_converted_data.extend(convert_to_documents([[], [meta]], meta['category']))

        # 파일 경로와 파일명을 결합하여 file_save_name_list에 있는지 확인
        if f"{meta['file_path']}/{meta['file_save_name']}" in file_save_name_list:
            filtered_files.add((meta["file_path"], meta["file_save_name"]))
            meta['text'] = document
            bm25_converted_data.extend(convert_to_documents([[], [meta]], meta['category']))

    # 파일명 필터 추가
    filter_conditions = [
        {"$and": [{"file_path": path}, {"file_save_name": name}]} for path, name in list(filtered_files)
    ]

    # 파일 경로 필터 추가
    # filtered_paths = list(filtered_paths)
    # if len(list(filtered_paths)) > 0:
    #     filter_conditions.append({"file_path": {"$in": list(filtered_paths)}})

    # category_pk 필터 추가
    if category_pk:
        filter_conditions.append({"category_pk": {"$in": category_pk}})

    # chroma 필터 조건.
    if len(filter_conditions) > 1:
        # 조건이 여러개라면 or로 묶는다.
        filter_dict = {"$or": filter_conditions}
    else:
        filter_dict = filter_conditions[0] if filter_conditions else {}

    # 리트리버 설정
    chroma_retriever = chroma_db.as_retriever(search_type="mmr", search_kwargs={"k": k, "fetch_k": fetch_k, "filter": filter_dict})
    bm25_retriever = BM25Retriever.from_documents(bm25_converted_data)
    ensenble_retriver = EnsembleRetriever(retrievers=[bm25_retriever, chroma_retriever], weights=[0.6, 0.4])

    # 리랭커 모델 로드
    model = HuggingFaceCrossEncoder(model_name=reranker_name)
    compressor = CrossEncoderReranker(model=model, top_n=top_n)

    # 문서 압축 검색기 초기화
    compression_retriever = ContextualCompressionRetriever(base_compressor=compressor, base_retriever=ensenble_retriver)
    return compression_retriever

def keyword_search(chroma_db, category_list=[], file_save_name_list=[], top_k=chunk_k):
    stored_metadata = chroma_db.get(include=["metadatas"])["metadatas"]
    text = chroma_db.get()['documents']

    # bm25용 데이터
    bm25_converted_data = []
    # chroma용 쿼리 필터
    filtered_paths, filtered_files = set(), set()

    for document, meta in zip(text, stored_metadata):
        # 파일 경로가 category_list에 포함되는지 확인
        if any(meta["file_path"].startswith(category) for category in category_list):
            filtered_paths.add(meta["file_path"])
            meta['text'] = document
            bm25_converted_data.extend(convert_to_documents([[], [meta]], meta['category']))

    # 파일 경로 필터 추가
    filter_conditions = []
    filtered_paths = list(filtered_paths)
    if len(list(filtered_paths)) > 0:
        filter_conditions.append({"file_path": {"$in": list(filtered_paths)}})

    # chroma 필터 조건.
    if len(filter_conditions) > 1:
        # 조건이 여러개라면 or로 묶는다.
        filter_dict = {"$or": filter_conditions}
    else:
        filter_dict = filter_conditions[0] if filter_conditions else {}

    # 리트리버 설정
    bm25_retriever = BM25Retriever.from_documents(bm25_converted_data)
    bm25_retriever.k = top_k 
#    tfidf_retriever = TFIDFRetriever.from_documents(bm25_converted_data)
#    tfidf_retriever.k = top_k 
#    ensenble_retriver = EnsembleRetriever(retrievers=[bm25_retriever, tfidf_retriever], weights=[0.5, 0.5])
    return bm25_retriever

# JSON 변환 함수
def convert_to_serializable(data):
    if isinstance(data, list):
        return [
            {
                "pageContent": doc.page_content,
                "page": doc.metadata.get("page_number"),
                "source": f"{doc.metadata.get('file_name')} {str(doc.metadata.get('page_number'))}page",
                "file_path": doc.metadata.get("file_path"),
                "file_save_name": doc.metadata.get("file_save_name"),
                "file_name": doc.metadata.get("file_name"),
                "chunk_number": doc.metadata.get("chunk_number"),
                "category_pk": doc.metadata.get("category_pk")
            }
            for doc in data
        ]
    return data


# 문서 삭제 함수
def delete_document_from_chroma(file_save_name: str):
    init_chroma()
    # 해당 문서가 존재하는지 확인
    stored_metadata = chroma_db.get(include=['metadatas'])
    target_indices = [
        i for i, meta in enumerate(stored_metadata['metadatas'])
        if meta.get("file_save_name") == file_save_name
    ]
    
    log(f"[delete_document_from_chroma] target_indices : {target_indices}")

    if not target_indices:
        return JSONResponse({"documentResult": {"result": False, "status_code": "error",
                                                "description": "Document not found in metadata"}}, status_code=404)

    # 해당 문서의 모든 청크 삭제
    chroma_db.delete(ids=[stored_metadata['ids'][i] for i in target_indices])

    return JSONResponse({"documentResult": {"result": True, "status_code": "success",
                                            "description": "Document vectors deleted successfully"}}, status_code=200)


def delete_document_from_chroma_by_session_id(session_id: str):
    init_chroma()
    stored_metadata = chroma_db.get(include=['metadatas'])
    target_indices = [
        i for i, meta in enumerate(stored_metadata['metadatas'])
        if meta.get("file_path") == session_id
    ]

    log(f"[delete_document_from_chroma_by_session_id] target_indices : {target_indices}")

    if not target_indices:
        return JSONResponse({"documentResult": {"result": False, "status_code": "error",
                                                "description": "Document not found in metadata"}}, status_code=404)

    # 해당 문서의 모든 청크 삭제
    chroma_db.delete(ids=[stored_metadata['ids'][i] for i in target_indices])

    return JSONResponse({"documentResult": {"result": True, "status_code": "success", "description": "delete success"}},
                        status_code=200)


# CUDA 메모리 초기화 함수
def clear_cuda_memory():
    try:
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
            torch.cuda.ipc_collect()
            log("[clear_cuda_memory] CUDA memory cleared successfully.")
    except Exception as e:
        log(f"[clear_cuda_memory] Error clearing CUDA memory: {str(e)}")


# 요청 바디를 위한 데이터 모델 정의
class DocumentRequest(BaseModel):
    kgptFileSaveName: str
    kgptFileName: str
    category: str = ""
    useMarker: bool = True
    categoryPk: str = ""


class MultipleMetadataRequest(BaseModel):
    prompt: str
    categoryList: List[str] = []
    kgptFileSaveNameList: List[str] = []
    topK: int = chunk_k
    useSearch: str = []
    categoryPk: List[str] = []

# 텍스트 임베딩을 위한 요청 모델
class TextEmbeddingRequest(BaseModel):
    content: str
    title: str


class MetricsInput(BaseModel):
    keyword1: str
    keyword2: str
    answer1: str
    answer2: str


class DeleteSessionRequest(BaseModel):
    sessionId: str


# 청킹된 데이터를 하나의 데이터로 만들기 위한 요청 모델(카테고리 내의 특정 문서)
class FullTextRequest(BaseModel):
    kgptFileSaveName: str
    category: str


# 청킹된 데이터를 하나의 데이터로 만들기 위한 요청 모델(카테고리 내의 5개 문서)
class FullTextListRequest(BaseModel):
    category: str


# 텍스트 문서를 처리하는 함수
def process_text_content(content: str, title: str):
    chunks = []
    chunk_metadata = []
    words = content.split()
    step = chunk_size - chunk_overlap
    text_chunks = [
        ' '.join(words[i:i + chunk_size])
        for i in range(0, len(words), step)
    ]
    for chunk_number, chunk in enumerate(text_chunks):
        chunks.append(chunk)
        chunk_metadata.append({
            "file_path": "text_contents",
            "file_save_name": f"{title}.txt",
            "file_name": title,
            "page_number": 1,
            "chunk_number": chunk_number + 1,
            "text": chunk
        })
    return chunks, chunk_metadata


# 전체 텍스트를 출력하는 함수
def process_raw_text(file_path: str, file_save_name: str, file_name: str):
    recursive = False
    parsed_text, chunk_metadata = extract_and_chunk_text_from_file(file_path, file_save_name, file_name, recursive)
    return parsed_text


def tokenize_korean(text: str) -> List[str]:
    text = text.replace('.', '').replace('?', '').replace('!', '')
    return text.split()


def calculate_all_metrics(
        keyword1: List[str],
        keyword2: List[str],
        answer1: str,
        answer2: str
) -> Dict[str, float]:
    all_keywords = list(set(keyword1 + keyword2))
    # 각 키워드 세트에 대해 이진 벡터 생성
    y_true = [1 if keyword in keyword1 else 0 for keyword in all_keywords]
    y_pred = [1 if keyword in keyword2 else 0 for keyword in all_keywords]

    try:
        # 키워드 기반 메트릭스 계산
        precision = float(precision_score(y_true, y_pred, zero_division=0))
        recall = float(recall_score(y_true, y_pred, zero_division=0))
        accuracy = float(accuracy_score(y_true, y_pred))
        f1 = float(f1_score(y_true, y_pred, zero_division=0))

        # 문장 토큰화
        tokenized_answer1 = tokenize_korean(answer1)
        tokenized_answer2 = tokenize_korean(answer2)
        log("[calculate_scores] tokenized_answer1 : ", tokenized_answer1)
        log("[calculate_scores] tokenized_answer2 : ", tokenized_answer2)

        # METEOR 점수 계산
        meteor = float(round(single_meteor_score(tokenized_answer1, tokenized_answer2), 4))

        log("[calculate_scores] precision : ", precision)
        log("[calculate_scores] recall : ", recall)
        log("[calculate_scores] accuracy : ", accuracy)
        log("[calculate_scores] f1 : ", f1)
        log("[calculate_scores] meteor : ", meteor)

        return {
            "precision": precision,
            "recall": recall,
            "accuracy": accuracy,
            "f1_score": f1,
            "meteor": meteor
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api-kgpt/grpc/calculate_scores")
async def get_metrics(data: MetricsInput) -> Dict[str, Union[float, str]]:
    # 실행 시간 측정 시작
    start_time = time.time()
    log(f"[calculate_scores_time_check Start] ######### time: {time.strftime("%H:%M:%S", time.localtime(start_time))} #########")

    try:
        log("[calculate_scores Start]")
        log("[calculate_scores] keyword1", data.keyword1)
        log("[calculate_scores] keyword2", data.keyword2)
        log("[calculate_scores] answer1", data.answer1)
        log("[calculate_scores] answer2", data.answer2)

        keyword1 = data.keyword1.split()
        keyword2 = data.keyword2.split()

        if not keyword1 or not keyword2 or not data.answer1 or not data.answer2:
            raise HTTPException(
                status_code=400,
                detail="keyworld or answer empty"
            )

        # 모든 성능 지표 계산
        metrics = calculate_all_metrics(keyword1, keyword2, data.answer1, data.answer2)
        log("[calculate_scores End]")

        # 종료
        end_time = time.time()
        log(f"[calculate_scores_time_check End] ######### time: {time.strftime("%H:%M:%S", time.localtime(end_time))} #########")
        elapsed_time = end_time - start_time
        log(f"[calculate_scores_time_check Total time] ######### Total: {elapsed_time:.2} seconds #########")

        return metrics

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# 파싱 전체 결과 출력 엔드포인트
@app.post("/api-kgpt/grpc/process-text")
async def process_raw_text_endpoint(request: DocumentRequest):
    """
    텍스트 문서를 처리하고 결과를 반환하는 API
    """
    kgpt_file_save_name = request.kgptFileSaveName
    kgpt_file_name = request.kgptFileName
    category = request.category
    log("[process-text] kgpt_file_save_name : ", kgpt_file_save_name)
    log("[process-text] kgpt_file_name : ", kgpt_file_name)
    log("[process-text] category : ", category)

    # 파일 경로 생성
    file_path = os.path.join(category, kgpt_file_save_name)
    log("[process-text] pdf_path : ", file_path)
    log("[process-text] os.path.exists(pdf_path)", os.path.exists(file_path))

    # 파일 존재 여부 확인
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail={
            "documentResult": {
                "result": False,
                "status_code": "error",
                "description": f"File not found: {file_path}"
            }
        })

    try:
        # 텍스트 처리 함수 호출
        parsed_text = process_raw_text(os.path.dirname(file_path), kgpt_file_save_name, kgpt_file_name)
        log('parsed_text', parsed_text)
        # 결과 반환
        return JSONResponse(content={
            "documentResult": {
                "result": True,
                "status_code": "success",
                "description": "Document processed successfully",
                "parsedText": parsed_text
            }
        })

    except Exception as e:
        log(f"[process_raw_text_endpoint] Error processing raw text: {str(e)}")
        raise HTTPException(status_code=500, detail={
            "documentResult": {
                "result": False,
                "status_code": "error",
                "description": f"Failed to process raw text: {str(e)}"
            }
        })


@app.post("/api-kgpt/grpc/text-embedder")
async def text_embedder(request: TextEmbeddingRequest):
    global chroma_db
    try:
        content = request.content
        title = request.title
        log(f"[text_embedder] title: {title}")
        # 텍스트 처리 및 청킹
        chunk_metadata = process_text_content(content, title)
        mark_documents = convert_to_documents(chunk_metadata, "onlyText")
        init_chroma()
        if chroma_db is None:
            chroma_db = Chroma.from_documents(
                mark_documents,
                embedding_model,
                persist_directory=DB_PATH,
                collection_name=COLLECTION_NAME
            )
        else:
            # 기존 DB에 추가
            chroma_db.add_documents(mark_documents)
        log(f"[text_embedder] success")
        return {
            "documentResult": {
                "result": True,
                "status_code": "success",
                "description": "Text content embedded successfully"
            }
        }
    except Exception as e:
        log(f"[text_embedder] Error embedding text content: {str(e)}")
        raise HTTPException(status_code=500, detail={
            "documentResult": {
                "result": False,
                "status_code": "error",
                "description": f"Failed to embed text content: {str(e)}"
            }
        })


# FastAPI 엔드포인트 정의(임베딩)
@app.post("/api-kgpt/grpc/document-embedder")
async def document_embedder(request: DocumentRequest):
    # 실행 시간 측정 시작
    start_time = time.time()
    log(f"[document_embedder_time_check Start] ######### time: {time.strftime("%H:%M:%S", time.localtime(start_time))} #########")

    kgpt_file_save_name = request.kgptFileSaveName
    kgpt_file_name = request.kgptFileName
    category = request.category
    use_marker = request.useMarker
    category_pk = request.categoryPk

    log("[document_embedder] kgpt_file_save_name ", kgpt_file_save_name)
    log("[document_embedder] kgpt_file_name ", kgpt_file_name)
    log("[document_embedder] category ", category)
    log("[document_embedder] use_marker ", use_marker)
    log("[document_embedder] category_pk ", category_pk)

    # 파일 경로 설정 및 존재 여부 확인
    pdf_path = os.path.join(category_pk, kgpt_file_save_name)
    log("[document_embedder] pdf_path ", pdf_path)
    log("[document_embedder] os.path.exists(pdf_path)", os.path.exists(pdf_path))
    if not os.path.exists(pdf_path):
        raise HTTPException(status_code=404, detail={
            "documentResult": {
                "result": False,
                "status_code": "error",
                "description": "File not found"
            }
        })

    # 파일 경로와 파일명을 load_or_create_embeddings 함수로 전달
    try:
        chroma_db = load_or_create_embeddings(
            file_path=os.path.dirname(pdf_path),  # 파일의 디렉토리 경로
            file_save_name=kgpt_file_save_name,
            file_name=kgpt_file_name,
            category=category,
            recursive=False,
            use_marker=use_marker,
            category_pk=category_pk
        )
        log(f"[document_embedder] Document '{kgpt_file_name}' successfully embedded in ChromaDB.")


    except Exception as e:
        log(f"[document_embedder] Error embedding document: {str(e)}")
        raise HTTPException(status_code=500, detail={
            "documentResult": {
                "result": False,
                "status_code": "error",
                "description": f"Failed to embed document: {str(e)}"
            }
        })

    # cuda 메모리 초기화
    clear_cuda_memory()

    # 종료
    end_time = time.time()
    log(f"[document_embedder_time_check End] ######### time: {time.strftime("%H:%M:%S", time.localtime(end_time))} #########")
    elapsed_time = end_time - start_time
    log(f"[document_embedder_time_check Total time] ######### Total: {elapsed_time:.2} seconds #########")

    return {
        "documentResult": {
            "result": True,
            "status_code": "success",
            "description": "Document embedded successfully"
        }
    }


# FastAPI 엔드포인트 정의, pdf를 마커 없이임베딩
@app.post("/api-kgpt/grpc/document-embedder-n-marker")
async def document_embedder_not_marker(request: DocumentRequest):
    kgpt_file_save_name = request.kgptFileSaveName
    kgpt_file_name = request.kgptFileName
    category = request.category

    log("[document_embedder] kgpt_file_save_name ", kgpt_file_save_name)
    log("[document_embedder] kgpt_file_name ", kgpt_file_name)
    log("[document_embedder] category ", category)

    # 파일 경로 설정 및 존재 여부 확인
    pdf_path = os.path.join(category, kgpt_file_save_name)
    log("[document_embedder] pdf_path ", pdf_path)
    log("[document_embedder] os.path.exists(pdf_path)", os.path.exists(pdf_path))
    if not os.path.exists(pdf_path):
        raise HTTPException(status_code=404, detail={
            "documentResult": {
                "result": False,
                "status_code": "error",
                "description": "File not found"
            }
        })

    # 파일 경로와 파일명을 load_or_create_embeddings 함수로 전달
    try:
        chroma_db = load_or_create_embeddings(
            file_path=os.path.dirname(pdf_path),  # 파일의 디렉토리 경로
            file_save_name=kgpt_file_save_name,
            file_name=kgpt_file_name,
            category=category,
            recursive=False,
            use_marker=False
        )
        log(f"[document_embedder] Document '{kgpt_file_name}' successfully embedded in ChromaDB.")


    except Exception as e:
        log(f"[document_embedder] Error embedding document: {str(e)}")
        raise HTTPException(status_code=500, detail={
            "documentResult": {
                "result": False,
                "status_code": "error",
                "description": f"Failed to embed document: {str(e)}"
            }
        })

    # cuda 메모리 초기화
    clear_cuda_memory()

    return {
        "documentResult": {
            "result": True,
            "status_code": "success",
            "description": "Document embedded successfully"
        }
    }


# FastAPI 엔드포인트 정의(임베딩) - splitter가 recursive
@app.post("/api-kgpt/grpc/document-embedder-re")
async def document_embedder_recursive(request: DocumentRequest):
    kgpt_file_save_name = request.kgptFileSaveName
    kgpt_file_name = request.kgptFileName
    category = request.category

    log("[document_embedder] kgpt_file_save_name ", kgpt_file_save_name)
    log("[document_embedder] kgpt_file_name ", kgpt_file_name)
    log("[document_embedder] category ", category)

    # 파일 경로 설정 및 존재 여부 확인
    pdf_path = os.path.join(category, kgpt_file_save_name)
    log("[document_embedder] pdf_path ", pdf_path)
    log("[document_embedder] os.path.exists(pdf_path)", os.path.exists(pdf_path))
    if not os.path.exists(pdf_path):
        raise HTTPException(status_code=404, detail={
            "documentResult": {
                "result": False,
                "status_code": "error",
                "description": "File not found"
            }
        })

    # 파일 경로와 파일명을 load_or_create_embeddings 함수로 전달
    try:
        chroma_db = load_or_create_embeddings(
            file_path=os.path.dirname(pdf_path),  # 파일의 디렉토리 경로
            file_save_name=kgpt_file_save_name,
            file_name=kgpt_file_name,
            category=category,
            recursive=True,
            use_marker=True
        )
        log(f"[document_embedder] Document '{kgpt_file_name}' successfully embedded in ChromaDB.")


    except Exception as e:
        log(f"[document_embedder] Error embedding document: {str(e)}")
        raise HTTPException(status_code=500, detail={
            "documentResult": {
                "result": False,
                "status_code": "error",
                "description": f"Failed to embed document: {str(e)}"
            }
        })

    # cuda 메모리 초기화
    clear_cuda_memory()

    return {
        "documentResult": {
            "result": True,
            "status_code": "success",
            "description": "Document embedded successfully"
        }
    }


# FastAPI 엔드포인트 정의
@app.post('/api-kgpt/grpc/document-delete')
async def document_delete(request_data: DocumentRequest):
    kgpt_file_save_name = request_data.kgptFileSaveName
    kgpt_file_name = request_data.kgptFileName
    category = request_data.category
    category_pk = request_data.categoryPk

    log("[document_delete] kgpt_file_save_name ", kgpt_file_save_name)
    log("[document_delete] kgpt_file_name ", kgpt_file_name)
    log("[document_delete] category ", category)
    log("[document_delete] category_pk ", category_pk)
    init_chroma()

    return delete_document_from_chroma(kgpt_file_save_name)


@app.post('/api-kgpt/grpc/delete_test_session')
async def delete_test_session(request_data: DeleteSessionRequest):
    session_id = request_data.sessionId
    log("[delete_test_session] session_id ", session_id)
    return delete_document_from_chroma_by_session_id(session_id)


# multiple_metadata FastAPI 엔드포인트 정의
@app.post("/api-kgpt/grpc/multiple_metadata")
async def multiple_metadata(request: MultipleMetadataRequest):
    # 실행 시간 측정 시작
    start_time = time.time()
    log(f"[document_retriever_time_check Start] ######### time: {time.strftime("%H:%M:%S", time.localtime(start_time))} #########")

    prompt = request.prompt
    category_list = request.categoryList
    file_save_name_list = request.kgptFileSaveNameList
    top_k = request.topK
    use_keyword = request.useSearch
    category_pk = request.categoryPk

    # 빈 프롬프트 체크
    if not prompt:
        raise HTTPException(status_code=404, detail="Prompt is empty")
    log("[multiple_metadata] prompt ", prompt)
    log("[multiple_metadata] categoryList ", category_list)
    log("[multiple_metadata] fileSaveNameList ", file_save_name_list)
    log("[multiple_metadata] top_k ", top_k)
    log("[multiple_metadata] use_keyword ", use_keyword)
    log("[multiple_metadata] category_pk ", category_pk)
    # ChromaDB 인스턴스 확인 및 retriever 생성
    init_chroma()
    if chroma_db is None:
        raise HTTPException(status_code=404, detail="ChromaDB instance not found")

    # Retriever 생성 및 Chain 생성
    try:
        # use_keyword가 "search"일 경우 키워드 기반 검색, 그렇지 않을 경우 기본 검색
        if use_keyword == "search":
            log("[multiple_metadata] Using keyword-based retriever")
            retriever = keyword_search(chroma_db, category_list=category_list, file_save_name_list=file_save_name_list, top_k=top_k)
        else:
            log("[multiple_metadata] Using default retriever")
            retriever = get_reranker(chroma_db, category_list=category_list, file_save_name_list=file_save_name_list, top_n=top_k, category_pk=category_pk)

    except Exception as e:
        log(f"[multiple_metadata] Error occurred while initializing the retriever: {e}")
        raise HTTPException(status_code=500, detail="Failed to initialize retriever.")

    try:
        # 프롬프트에 대해 chain 실행하여 검색된 문서 가져오기
        retrieved_docs = retriever.invoke(prompt)
        log("[generate_question_from_chunks] chunk_k ", len(retrieved_docs))

        # 문서 필터링: category_list에 포함된 파일명과 일치하는 것만 필터링
        # relevant_chunks = [
        #    doc for doc in retrieved_docs
        #    if doc.metadata.get('file_path') in category_list]
        # log("최종 Chroma_result 개수 : ", len(relevant_chunks))

        # 사용할 문서 선택 (필터링된 청크가 있으면 사용하고, 없으면 원본 사용)
        # response_chunks = relevant_chunks if len(relevant_chunks) > 0 else retrieved_docs

        # 응답 데이터 포맷팅
        response = convert_to_serializable(retrieved_docs)
        if not response:
            log("[multiple_metadata] No relevant documents found after filtering")
        log("[RESPONSE] >>>> >>>> >>>> >>> \n", json.dumps(response, indent=2, ensure_ascii=False))
        log("=" * 124)
        log("=" * 124)

        # 종료
        end_time = time.time()
        log(f"[document_retriever_time_check End] ######### time: {time.strftime("%H:%M:%S", time.localtime(end_time))} #########")
        elapsed_time = end_time - start_time
        log(f"[document_retriever_time_check Total time] ######### Total: {elapsed_time:.2} seconds #########")

        return JSONResponse(content=response)

    except Exception as e:
        log(f"[multiple_metadata] Error retrieving relevant chunks: {e}")
        raise HTTPException(status_code=500, detail="Failed to retrieve relevant chunks")


# 특정 문서 하나의 청크 데이터를 하나의 텍스트로 가져오는 FastAPI 엔드포인트 정의
@app.post("/api-kgpt/grpc/full_text_metadata")
async def full_text_metadata(request: FullTextRequest):
    kgpt_file_save_name = request.kgptFileSaveName
    category = request.category

    log("[full_text_metadata] kgptFileSaveName ", kgpt_file_save_name)
    log("[full_text_metadata] category ", category)

    init_chroma()
    if chroma_db is None:
        raise HTTPException(status_code=404, detail="ChromaDB instance not found")

    try:
        response = get_full_chunk_text(chroma_db, category, kgpt_file_save_name)
        return JSONResponse(content=response)

    except Exception as e:
        log(f"[full_text_metadata] Error retrieving relevant chunks: {e}")
        raise HTTPException(status_code=500, detail="Failed to find file")


# 카테고리에서 문서 5개를 통으로 리스트에 가져오는 FastAPI 엔드포인트 정의
@app.post("/api-kgpt/grpc/full_text_multiple_metadata")
async def full_text_multiple_metadata(request: FullTextListRequest):
    category = request.category

    log("[full_text_multiple_metadata] category ", category)

    init_chroma()
    if chroma_db is None:
        raise HTTPException(status_code=404, detail="ChromaDB instance not found")

    try:
        response = get_full_chunk_text_list(chroma_db, category)
        return JSONResponse(content=response)

    except Exception as e:
        log(f"[full_text_multiple_metadata] Error retrieving relevant chunks: {e}")
        raise HTTPException(status_code=500, detail="Failed to find file")
    
    
if __name__ == "__main__":
    nest_asyncio.apply()
    uvicorn.run(app, host="0.0.0.0", port=5000, reload=False)